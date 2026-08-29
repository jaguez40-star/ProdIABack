"""
Prototipo de ingesta — Robustez V2.0
====================================
Carga UN archivo .xlsm del Reporte Diario de Producción a la BD `robustez` (PostgreSQL).

Estrategia D3 (degradación elegante): se ingiere todo archivo y se preserva lo que trae.
- Detecta capa raw por presencia de hoja (BDP_datos_*) → marca config_reporte.
- Bronze: BDP_datos_dia/mes/programa (tipadas) + demás hojas a hoja_landing (JSONB).
- core: dims (fuente, socio, concepto, ...) + facts ECP (dia/mes/programa) + comentarios.
- Idempotencia: upsert "última gana" (ON CONFLICT) en facts; delete-by-reporte en bronze/comentarios.

Uso:
  uv run --with sqlalchemy --with "psycopg[binary]" --with openpyxl --with python-dotenv \
      python etl/ingesta_prototipo.py "Doc_Desing/20241004_Reporte New Diario de Producción.xlsm"
"""
from __future__ import annotations
import os, sys, re, json, time, datetime as dt
from pathlib import Path
from openpyxl import load_workbook
import sqlalchemy as sa
from dotenv import load_dotenv

CHUNK = 10_000
NOISE = {"", "#REF!", "#DIV/0!", "#N/A", "#VALUE!", "#NAME?", "(en blanco)", "(EN BLANCO)"}

# ---------------------------------------------------------------- helpers parseo
def s(v):
    if v is None: return None
    t = str(v).strip()
    return None if t in NOISE else t

def num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return v
    t = str(v).strip()
    if t in NOISE: return None
    try: return float(t)
    except ValueError: return None

def to_date(v):
    if v is None or v == "" or v == 0: return None
    if isinstance(v, dt.datetime): return v.date()
    if isinstance(v, dt.date): return v
    t = str(v).strip()
    if t in NOISE: return None
    t = t.split(".")[0]          # "20240930.0" -> "20240930"
    if re.fullmatch(r"\d{8}", t):
        try: return dt.date(int(t[:4]), int(t[4:6]), int(t[6:8]))
        except ValueError: return None
    try: return dt.date.fromisoformat(t[:10])
    except ValueError: return None

# ---------------------------------------------------------------- dims pequeñas (cache + upsert)
class DimCache:
    """Resuelve nombre->id de una dim con UNIQUE(nombre/codigo); upsert si falta."""
    def __init__(self, conn, table, idcol, namecol):
        self.conn, self.table, self.idcol, self.namecol = conn, table, idcol, namecol
        self.cache = {}
        for r in conn.execute(sa.text(f"SELECT {idcol}, {namecol} FROM {table}")):
            self.cache[r[1]] = r[0]
    def get(self, name, extra_cols=""):
        if name is None: return None
        if name in self.cache: return self.cache[name]
        rid = self.conn.execute(sa.text(
            f"INSERT INTO {self.table} ({self.namecol}{extra_cols}) VALUES (:n{',NULL'*extra_cols.count(',')}) "
            f"ON CONFLICT ({self.namecol}) DO UPDATE SET {self.namecol}=EXCLUDED.{self.namecol} "
            f"RETURNING {self.idcol}"), {"n": name}).scalar()
        self.cache[name] = rid
        return rid

# ---------------------------------------------------------------- dim_fuente / dim_fecha
FUENTE_COLS = ["nombre","campo","contrato","tipo_contrato","operador","modalidad","operacion",
               "nacionalidad","gerencia","grupo1","grupo2","grupo3","activos","fuente_contrato"]

def upsert_fuentes(conn, fuentes: dict, reporte_id: int):
    if not fuentes: return
    sets = ", ".join(f"{c}=COALESCE(EXCLUDED.{c}, core.dim_fuente.{c})" for c in FUENTE_COLS)
    cols = ", ".join(FUENTE_COLS)
    binds = ", ".join(f":{c}" for c in FUENTE_COLS)
    stmt = sa.text(f"""
        INSERT INTO core.dim_fuente (fuente_id, {cols}, reporte_id_origen)
        VALUES (:fuente_id, {binds}, :rep)
        ON CONFLICT (fuente_id) DO UPDATE SET {sets},
            reporte_id_origen = EXCLUDED.reporte_id_origen, updated_at = now()""")
    rows = []
    for fid, a in fuentes.items():
        d = {"fuente_id": fid, "rep": reporte_id}
        for c in FUENTE_COLS: d[c] = a.get(c)
        rows.append(d)
    conn.execute(stmt, rows)

def ensure_fechas(conn, fechas: set):
    rows = []
    for f in fechas:
        if f is None: continue
        rows.append({"f": f, "a": f.year, "m": f.month, "d": f.day,
                     "t": (f.month-1)//3+1, "dw": f.isoweekday(), "sm": (f.day-1)//7+1})
    if rows:
        conn.execute(sa.text("""
            INSERT INTO core.dim_fecha (fecha,anio,mes,dia,trimestre,dia_semana,semana_mes)
            VALUES (:f,:a,:m,:d,:t,:dw,:sm) ON CONFLICT (fecha) DO NOTHING"""), rows)

# ---------------------------------------------------------------- detección + config
def detectar(wb):
    sheets = set(wb.sheetnames)
    tiene_raw = {"BDP_datos_dia", "BDP_datos_mes", "BDP_Programa"} <= sheets
    return tiene_raw

def get_reporte(conn, path: Path, tiene_raw: bool):
    m = re.search(r"(\d{8})", path.name)
    fecha_rep = to_date(m.group(1)) if m else None
    tipo = "NEW" if tiene_raw else "STD"
    nivel = "FULL" if tiene_raw else "SIN_ECP"
    rid = conn.execute(sa.text("""
        INSERT INTO core.config_reporte (fecha_reporte, archivo_nombre, tipo_archivo, tiene_raw, nivel_detalle)
        VALUES (:fr, :an, :tp, :raw, :nv)
        ON CONFLICT (fecha_reporte) DO UPDATE SET
            archivo_nombre=EXCLUDED.archivo_nombre, tipo_archivo=EXCLUDED.tipo_archivo,
            tiene_raw=EXCLUDED.tiene_raw, nivel_detalle=EXCLUDED.nivel_detalle, ingested_at=now()
        RETURNING reporte_id""",
        ), {"fr": fecha_rep, "an": path.name, "tp": tipo, "raw": tiene_raw, "nv": nivel}).scalar()
    return rid, fecha_rep

def log(conn, reporte_id, hoja, destino, leidas, ins, estado="OK", msg=None):
    conn.execute(sa.text("""
        INSERT INTO core.ingesta_log (reporte_id,hoja,tabla_destino,filas_leidas,filas_insertadas,estado,mensaje)
        VALUES (:r,:h,:d,:l,:i,:e,:m)"""),
        {"r": reporte_id, "h": hoja, "d": destino, "l": leidas, "i": ins, "e": estado, "m": msg})

# ---------------------------------------------------------------- BRONZE
def land_bronze_typed(conn, ws, table, cols, reporte_id):
    """Inserta una hoja plana raw en su tabla bronze tipada (todo TEXT). Pasada por chunks."""
    conn.execute(sa.text(f"DELETE FROM bronze.{table} WHERE reporte_id=:r"), {"r": reporte_id})
    binds = ", ".join(f":{c}" for c in cols)
    stmt = sa.text(f"INSERT INTO bronze.{table} (reporte_id, fila_origen, {', '.join(cols)}) "
                   f"VALUES (:_rep, :_fila, {binds})")
    buf, total = [], 0
    it = ws.iter_rows(values_only=True); next(it, None)  # salta encabezado
    for i, row in enumerate(it, start=2):
        if row is None or all(c is None for c in row): continue
        d = {"_rep": reporte_id, "_fila": i}
        for j, c in enumerate(cols): d[c] = (None if j >= len(row) or row[j] is None else str(row[j]))
        buf.append(d); total += 1
        if len(buf) >= CHUNK: conn.execute(stmt, buf); buf.clear()
    if buf: conn.execute(stmt, buf)
    return total

def land_landing(conn, ws, hoja, reporte_id):
    """Aterriza cualquier otra hoja a bronze.hoja_landing (JSONB), preservando celdas."""
    conn.execute(sa.text("DELETE FROM bronze.hoja_landing WHERE reporte_id=:r AND hoja=:h"),
                 {"r": reporte_id, "h": hoja})
    it = ws.iter_rows(values_only=True)
    header = next(it, None) or ()
    keys = [str(h) if h is not None else f"col{j}" for j, h in enumerate(header)]
    stmt = sa.text("""INSERT INTO bronze.hoja_landing (reporte_id,hoja,fila_origen,payload)
                      VALUES (:r,:h,:f, CAST(:p AS jsonb))""")
    buf, total = [], 0
    for i, row in enumerate(it, start=2):
        if row is None or all(c is None for c in row): continue
        payload = {keys[j] if j < len(keys) else f"col{j}":
                   (None if v is None else str(v)) for j, v in enumerate(row)}
        buf.append({"r": reporte_id, "h": hoja, "f": i, "p": json.dumps(payload, ensure_ascii=False)})
        total += 1
        if len(buf) >= CHUNK: conn.execute(stmt, buf); buf.clear()
    if buf: conn.execute(stmt, buf)
    return total

# ---------------------------------------------------------------- FACTS ECP
def load_fact_dia(conn, ws, reporte_id, caches):
    DIA = sa.text("""
      INSERT INTO core.fact_produccion_dia_ecp
        (fecha,fuente_id,vice_id,socio_id,concepto_id,tipo_producto_id,producto,grupo_prod,
         propietario,volumen,porcentaje,voldismez,vol_estimado,promedio,reporte_id)
      VALUES (:fecha,:fuente_id,:vice_id,:socio_id,:concepto_id,:tipo_producto_id,:producto,
              :grupo_prod,:propietario,:volumen,:porcentaje,:voldismez,:vol_estimado,:promedio,:rep)
      ON CONFLICT ON CONSTRAINT uk_dia DO UPDATE SET
        volumen=EXCLUDED.volumen, porcentaje=EXCLUDED.porcentaje, voldismez=EXCLUDED.voldismez,
        vol_estimado=EXCLUDED.vol_estimado, promedio=EXCLUDED.promedio, reporte_id=EXCLUDED.reporte_id""")
    vice, socio, conc, tipo = caches
    buf, fuentes, fechas, total, skip = [], {}, set(), 0, 0
    it = ws.iter_rows(values_only=True); next(it, None)
    for row in it:
        if row is None or row[6] is None: continue
        fecha = to_date(row[8]); idbdp = num(row[6])
        if fecha is None or idbdp is None: skip += 1; continue
        fid = int(idbdp)
        fuentes[fid] = {"nombre": s(row[5]), "contrato": s(row[4]), "tipo_contrato": s(row[3]),
                        "operador": s(row[2]), "modalidad": s(row[17]), "operacion": s(row[18]),
                        "nacionalidad": s(row[19]), "gerencia": s(row[16]), "grupo1": s(row[20]),
                        "grupo2": s(row[21]), "grupo3": s(row[22]), "activos": s(row[26]),
                        "fuente_contrato": s(row[7])}
        fechas.add(fecha)
        buf.append({"fecha": fecha, "fuente_id": fid, "vice_id": vice.get(s(row[25])),
                    "socio_id": socio.get(s(row[1])), "concepto_id": conc.get(s(row[0])),
                    "tipo_producto_id": tipo.get(s(row[15])), "producto": s(row[14]) or "",
                    "grupo_prod": s(row[13]) or "", "propietario": s(row[12]) or "",
                    "volumen": num(row[23]), "porcentaje": num(row[24]), "voldismez": num(row[27]),
                    "vol_estimado": num(row[28]), "promedio": num(row[29]), "rep": reporte_id})
        total += 1
        if len(buf) >= CHUNK:
            ensure_fechas(conn, fechas); upsert_fuentes(conn, fuentes, reporte_id)
            conn.execute(DIA, buf); buf.clear(); fuentes.clear(); fechas.clear()
    if buf:
        ensure_fechas(conn, fechas); upsert_fuentes(conn, fuentes, reporte_id); conn.execute(DIA, buf)
    return total, skip

def load_fact_mes(conn, ws, reporte_id, caches):
    MES = sa.text("""
      INSERT INTO core.fact_produccion_mes_ecp
        (fecha,fuente_id,vice_id,socio_id,concepto_id,tipo_producto_id,producto,escenario_id,
         proceso_id,grupo_prod,negocio,volumen,porcentaje,voldismez,bpd_m,bpda_ac,bpd_a,
         bpdeq_m,blseq,bpdeq_a,reporte_id)
      VALUES (:fecha,:fuente_id,:vice_id,:socio_id,:concepto_id,:tipo_producto_id,:producto,
              :escenario_id,:proceso_id,:grupo_prod,:negocio,:volumen,:porcentaje,:voldismez,
              :bpd_m,:bpda_ac,:bpd_a,:bpdeq_m,:blseq,:bpdeq_a,:rep)
      ON CONFLICT ON CONSTRAINT uk_mes DO UPDATE SET
        volumen=EXCLUDED.volumen, porcentaje=EXCLUDED.porcentaje, voldismez=EXCLUDED.voldismez,
        bpd_m=EXCLUDED.bpd_m, bpda_ac=EXCLUDED.bpda_ac, bpd_a=EXCLUDED.bpd_a, bpdeq_m=EXCLUDED.bpdeq_m,
        blseq=EXCLUDED.blseq, bpdeq_a=EXCLUDED.bpdeq_a, negocio=EXCLUDED.negocio, reporte_id=EXCLUDED.reporte_id""")
    vice, socio, conc, tipo, esc, proc = caches
    buf, fuentes, fechas, total, skip = [], {}, set(), 0, 0
    it = ws.iter_rows(values_only=True); next(it, None)
    for row in it:
        if row is None or row[8] is None: continue
        fecha = to_date(row[11]); idbdp = num(row[8])
        if fecha is None or idbdp is None: skip += 1; continue
        fid = int(idbdp)
        fuentes[fid] = {"nombre": s(row[7]), "contrato": s(row[6]), "tipo_contrato": s(row[5]),
                        "operador": s(row[3]), "modalidad": s(row[33]), "operacion": s(row[34]),
                        "nacionalidad": s(row[36]), "gerencia": s(row[32]), "grupo1": s(row[37]),
                        "grupo2": s(row[38]), "grupo3": s(row[39]), "activos": s(row[58]),
                        "fuente_contrato": s(row[10])}
        fechas.add(fecha)
        buf.append({"fecha": fecha, "fuente_id": fid, "vice_id": vice.get(s(row[57])),
                    "socio_id": socio.get(s(row[1])), "concepto_id": conc.get(s(row[0])),
                    "tipo_producto_id": tipo.get(s(row[20])), "producto": s(row[19]) or "",
                    "escenario_id": esc.get(s(row[14])), "proceso_id": proc.get(s(row[15])),
                    "grupo_prod": s(row[18]) or "", "negocio": s(row[21]),
                    "volumen": num(row[40]), "porcentaje": num(row[41]), "voldismez": num(row[42]),
                    "bpd_m": num(row[43]), "bpda_ac": num(row[44]), "bpd_a": num(row[46]),
                    "bpdeq_m": num(row[47]), "blseq": num(row[48]), "bpdeq_a": num(row[49]), "rep": reporte_id})
        total += 1
        if len(buf) >= CHUNK:
            ensure_fechas(conn, fechas); upsert_fuentes(conn, fuentes, reporte_id)
            conn.execute(MES, buf); buf.clear(); fuentes.clear(); fechas.clear()
            print(f"    ... mes {total:,} filas", flush=True)
    if buf:
        ensure_fechas(conn, fechas); upsert_fuentes(conn, fuentes, reporte_id); conn.execute(MES, buf)
    return total, skip

def load_fact_programa(conn, ws, reporte_id, caches):
    PRG = sa.text("""
      INSERT INTO core.fact_programa_ecp
        (fecha,vice_id,tipo_producto_id,fuente_id,area,campo,version,fecha_version,
         volumen,produccion_total,part_ecp,reporte_id)
      VALUES (:fecha,:vice_id,:tipo_producto_id,:fuente_id,:area,:campo,:version,:fecha_version,
              :volumen,:produccion_total,:part_ecp,:rep)
      ON CONFLICT ON CONSTRAINT uk_prog DO UPDATE SET
        volumen=EXCLUDED.volumen, produccion_total=EXCLUDED.produccion_total,
        part_ecp=EXCLUDED.part_ecp, fecha_version=EXCLUDED.fecha_version, reporte_id=EXCLUDED.reporte_id""")
    vice, tipo = caches
    buf, fechas, total, skip = [], set(), 0, 0
    it = ws.iter_rows(values_only=True); next(it, None)
    for row in it:
        if row is None or row[0] is None: continue
        fecha = to_date(row[0])
        if fecha is None: skip += 1; continue
        idbdp = num(row[10]); fid = int(idbdp) if idbdp is not None else None
        fechas.add(fecha)
        buf.append({"fecha": fecha, "vice_id": vice.get(s(row[1])),
                    "tipo_producto_id": tipo.get(s(row[8])), "fuente_id": fid,
                    "area": s(row[9]) or "", "campo": s(row[7]) or "", "version": s(row[3]) or "",
                    "fecha_version": to_date(row[4]), "volumen": num(row[6]),
                    "produccion_total": num(row[12]), "part_ecp": num(row[13]), "rep": reporte_id})
        total += 1
        if len(buf) >= CHUNK:
            ensure_fechas(conn, fechas); conn.execute(PRG, buf); buf.clear(); fechas.clear()
    if buf:
        ensure_fechas(conn, fechas); conn.execute(PRG, buf)
    # NOTA: fuente_id de programa puede referenciar IDBDP no presente en dia/mes; si la FK falla,
    # esos IDBDP deben sembrarse en dim_fuente. Se maneja abajo en main (pre-siembra).
    return total, skip

# ---------------------------------------------------------------- COMENTARIOS
def load_comentarios(conn, ws, reporte_id, tipo_cache):
    conn.execute(sa.text("DELETE FROM core.fact_comentarios_produccion WHERE reporte_id=:r"), {"r": reporte_id})
    stmt = sa.text("""INSERT INTO core.fact_comentarios_produccion
                      (tipo_producto_id,activos,area,comentario,reporte_id)
                      VALUES (:t,:a,:ar,:c,:r)""")
    buf, total = [], 0
    it = ws.iter_rows(values_only=True); next(it, None)
    for row in it:
        if row is None or len(row) < 4: continue
        coment = s(row[3]) or (s(row[4]) if len(row) > 4 else None)
        if not coment or coment == "0": continue
        buf.append({"t": tipo_cache.get(s(row[0])), "a": s(row[1]), "ar": s(row[2]),
                    "c": coment, "r": reporte_id})
        total += 1
    if buf: conn.execute(stmt, buf)
    return total

# ---------------------------------------------------------------- bronze column specs
BZ_DIA = ["concepto","socio","operador","tipocontrato","contrato","fuente","idbdp","fuentecontrato",
          "fecha","mes","anio","escenario","propietario","grupoprod","producto","tipoproducto",
          "gerencia","modalidad","operacion","nacionalidad","grupo1","grupo2","grupo3","volumen",
          "porcentaje","vice","activos","voldismez","vol_estimado","promedio"]
BZ_MES = ["concepto","socio","ba_id","operador","grupooperador","tipocontrato","contrato","fuente",
          "idbdp","fuentecontratoplot","fuentecontrato","fecha","mes","anio","escenario","proceso",
          "row_changed_by","propietario","grupoprod","producto","tipoproducto","negocio","nuevagerencia",
          "superintendencia","tag","tipofuente","tagdescripcion","fechaefprop","fechaexprop","fechaefpden",
          "fechaexpden","negociovpr","gerencia","modalidad","operacion","tipocrudo","nacionalidad","grupo1",
          "grupo2","grupo3","volumen","porcentaje","voldismez","bpd_m","bpda_ac","bpdac_5","bpd_a","bpdeq_m",
          "blseq","bpdeq_a","nodo","diluyente","linea_estrategica","mezcla_siv_gas","producto_yacimiento",
          "proyeccion","esc_proy","vice","activos"]
BZ_PRG = ["fecha","vice","gerencia","version","fecha_version","estado","volumen","campo","producto",
          "area","idbdp","contrato","produccion_total","part_ecp"]

RAW_SHEETS = {"BDP_datos_dia", "BDP_datos_mes", "BDP_Programa"}

# ---------------------------------------------------------------- FILIALES / POP / PROMEDIOS / CONFIG
EMP_NORM = {"EAI": "America", "EA": "America", "AMERICA": "America",
            "HOCOL": "Hocol", "PERMIAN": "Permian"}
PROD_NORM = {"CRUDO": "CRUDO", "GAS": "GAS", "BLANCOS": "BLANCOS", "BLANCO": "BLANCOS"}

def norm_emp(e):
    if e is None: return None
    return EMP_NORM.get(e.strip().upper(), e.strip())

def norm_prod(p):
    if p is None: return None
    return PROD_NORM.get(p.strip().upper())

def split_label(lbl):
    """'Hocol (crudo)' -> ('Hocol','crudo')."""
    m = re.match(r"^\s*(.+?)\s*\(\s*([^)]+?)\s*\)?\s*$", lbl)
    return (m.group(1), m.group(2)) if m else (None, None)

def load_filiales(conn, ws, reporte_id, emp, prod, treg):
    """Bloques REAL y PROGRAMA de 'Producción filiales' → fact_produccion_diaria (unpivot)."""
    stmt = sa.text("""
      INSERT INTO core.fact_produccion_diaria (empresa_id,producto_id,tipo_id,fecha,valor_produccion,reporte_id)
      VALUES (:e,:p,:t,:f,:v,:r)
      ON CONFLICT ON CONSTRAINT uk_fil DO UPDATE SET
        valor_produccion=EXCLUDED.valor_produccion, reporte_id=EXCLUDED.reporte_id""")
    tipo, dates, buf, fechas, total = None, [], [], set(), 0
    for row in ws.iter_rows(values_only=True):
        a = s(row[0]) if row else None
        if a is None: continue
        au = a.upper()
        if au == "REAL": tipo, dates = "Real", []; continue
        if au == "PROGRAMA": tipo, dates = "Programa", []; continue
        if au.startswith("PROYEC"): tipo, dates = None, []; continue
        if au == "EMPRESA": dates = [to_date(v) for v in row[1:]]; continue
        if tipo is None or au.startswith("TOTAL"): continue
        emp_n, prod_n = split_label(a)
        eid, pid, tid = emp.get(norm_emp(emp_n)), prod.get(norm_prod(prod_n)), treg.get(tipo)
        if not (eid and pid and tid): continue
        for j, val in enumerate(row[1:]):
            f = dates[j] if j < len(dates) else None
            v = num(val)
            if f is None or v is None: continue
            fechas.add(f)
            buf.append({"e": eid, "p": pid, "t": tid, "f": f, "v": v, "r": reporte_id})
    ensure_fechas(conn, fechas)
    if buf: conn.execute(stmt, buf)
    return len(buf)

def load_pop(conn, ws, reporte_id, emp):
    """'POP Filiales y Exploración' filas TOTAL <empresa> → fact_plan_mensual (pop_kbd)."""
    stmt = sa.text("""
      INSERT INTO core.fact_plan_mensual (empresa_id,anio,mes,pop_kbd,segmento,reporte_id)
      VALUES (:e,:a,:m,:p,'Filiales',:r)
      ON CONFLICT ON CONSTRAINT uk_plan DO UPDATE SET pop_kbd=EXCLUDED.pop_kbd, reporte_id=EXCLUDED.reporte_id""")
    dates, buf = [], []
    for row in ws.iter_rows(values_only=True):
        if row is None: continue
        if s(row[1]) == "Producto":                       # fila de encabezado con fechas
            dates = [to_date(v) for v in row]
            continue
        b = s(row[1])
        emp_n = norm_emp(s(row[2]))
        if not (b and b.upper().startswith("TOTAL") and emp_n in ("Hocol", "America", "Permian")):
            continue
        eid = emp.get(emp_n)
        for j, val in enumerate(row):
            f = dates[j] if j < len(dates) else None
            v = num(val)
            if f is None or v is None: continue
            buf.append({"e": eid, "a": f.year, "m": f.month, "p": v/1000.0, "r": reporte_id})
    if buf: conn.execute(stmt, buf)
    return len(buf)

def load_promedios(conn, ws, reporte_id, emp, prod):
    """Sección 'REAL PROMEDIO MES (YTD)' de INICIO → fact_promedio_validado."""
    stmt = sa.text("""
      INSERT INTO core.fact_promedio_validado (empresa_id,producto_id,anio,mes,promedio_validado,reporte_id)
      VALUES (:e,:p,:a,:m,:v,:r)
      ON CONFLICT ON CONSTRAINT uk_prom DO UPDATE SET
        promedio_validado=EXCLUDED.promedio_validado, reporte_id=EXCLUDED.reporte_id""")
    dates, buf, in_sec = [], [], False
    for row in ws.iter_rows(values_only=True):
        a = s(row[0]) if row else None
        if a == "Producto" and s(row[1]) == "Empresa":
            dates = [to_date(v) for v in row]; in_sec = True; continue
        if not in_sec: continue
        if a is None or a.upper() == "TOTAL": continue
        pid = prod.get(norm_prod(a)); eid = emp.get(norm_emp(s(row[1])))
        if not (pid and eid): continue
        for j, val in enumerate(row):
            f = dates[j] if j < len(dates) else None
            v = num(val)
            if f is None or v is None: continue
            buf.append({"e": eid, "p": pid, "a": f.year, "m": f.month, "v": v, "r": reporte_id})
    if buf: conn.execute(stmt, buf)
    return len(buf)

def update_config_inicio(conn, ws, reporte_id):
    """Completa fecha_corte/mes_inicio/mes_fin/version_semana/anio_inicio/dias_anio desde INICIO."""
    lab = {}
    for row in ws.iter_rows(min_row=1, max_row=40, max_col=4, values_only=True):
        b = s(row[1])
        if b: lab[b] = row[2]
    def n(x): v = num(x); return int(v) if v is not None else None
    vals = {
        "fc": to_date(lab.get("Día de corte")),
        "mi": to_date(lab.get("1er dia del mes")),
        "mf": to_date(lab.get("MES CORTE")),
        "vs": n(lab.get("Version Semana")),
        "ai": (to_date(lab.get("Fecha inicial año")).year if to_date(lab.get("Fecha inicial año")) else None),
        "da": n(lab.get("Días del año")),
        "r": reporte_id}
    conn.execute(sa.text("""
        UPDATE core.config_reporte SET fecha_corte=:fc, mes_inicio=:mi, mes_fin=:mf,
            version_semana=:vs, anio_inicio=:ai, dias_anio=:da WHERE reporte_id=:r"""), vals)
    return vals

# ---------------------------------------------------------------- MAIN
def main():
    load_dotenv(Path(__file__).resolve().parents[1] / ".env")
    url = os.environ["DATABASE_URL"]
    argv = [a for a in sys.argv[1:] if not a.startswith("--")]
    solo_deriv = "--solo-derivadas" in sys.argv      # salta bronze+ECP (iteración rápida)
    path = Path(argv[0] if argv else
                r"Doc_Desing/20241004_Reporte New Diario de Producción.xlsm").resolve()
    print(f"Archivo: {path.name}" + ("  [solo-derivadas]" if solo_deriv else ""))
    t0 = time.time()
    wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    tiene_raw = detectar(wb)
    print(f"Tipo: {'NEW (con raw)' if tiene_raw else 'STD (sin raw)'}  | hojas={len(wb.sheetnames)}")

    engine = sa.create_engine(url, future=True)
    with engine.begin() as conn:
        reporte_id, fecha_rep = get_reporte(conn, path, tiene_raw)
        print(f"reporte_id={reporte_id}  fecha_reporte={fecha_rep}")

        # caches de dims
        vice  = DimCache(conn, "core.dim_vicepresidencia", "vice_id", "codigo")
        socio = DimCache(conn, "core.dim_socio", "socio_id", "nombre")
        conc  = DimCache(conn, "core.dim_concepto", "concepto_id", "nombre")
        tipo  = DimCache(conn, "core.dim_tipo_producto", "tipo_producto_id", "nombre")
        esc   = DimCache(conn, "core.dim_escenario", "escenario_id", "nombre")
        proc  = DimCache(conn, "core.dim_proceso", "proceso_id", "nombre")
        emp   = DimCache(conn, "core.dim_empresa", "empresa_id", "nombre")
        treg  = DimCache(conn, "core.dim_tipo_registro", "tipo_id", "nombre")

        # ---- BRONZE typed (raw) ----
        if tiene_raw and not solo_deriv:
            for sheet, table, cols in [("BDP_datos_dia", "bdp_datos_dia", BZ_DIA),
                                       ("BDP_datos_mes", "bdp_datos_mes", BZ_MES),
                                       ("BDP_Programa",  "bdp_programa",  BZ_PRG)]:
                n = land_bronze_typed(conn, wb[sheet], table, cols, reporte_id)
                log(conn, reporte_id, sheet, f"bronze.{table}", n, n)
                print(f"  bronze.{table}: {n:,} filas")

        # ---- BRONZE landing (demás hojas) ----
        if not solo_deriv:
            for sheet in wb.sheetnames:
                if sheet in RAW_SHEETS: continue
                n = land_landing(conn, wb[sheet], sheet, reporte_id)
                log(conn, reporte_id, sheet, "bronze.hoja_landing", n, n)
            print(f"  bronze.hoja_landing: {len(wb.sheetnames)-(3 if tiene_raw else 0)} hojas aterrizadas")

        # ---- FACTS ECP (solo si raw) ----
        if tiene_raw and not solo_deriv:
            n, sk = load_fact_dia(conn, wb["BDP_datos_dia"], reporte_id, (vice, socio, conc, tipo))
            log(conn, reporte_id, "BDP_datos_dia", "core.fact_produccion_dia_ecp", n+sk, n)
            print(f"  fact_dia: {n:,} filas ({sk} descartadas)")

            n, sk = load_fact_mes(conn, wb["BDP_datos_mes"], reporte_id, (vice, socio, conc, tipo, esc, proc))
            log(conn, reporte_id, "BDP_datos_mes", "core.fact_produccion_mes_ecp", n+sk, n)
            print(f"  fact_mes: {n:,} filas ({sk} descartadas)")

            # pre-siembra de fuentes de programa que no estén aún (IDBDP nuevos) para no violar FK
            prg = wb["BDP_Programa"]
            it = prg.iter_rows(values_only=True); next(it, None)
            extra = {}
            for r in it:
                idb = num(r[10]) if r and len(r) > 10 else None
                if idb is not None: extra[int(idb)] = {"campo": s(r[7]), "grupo1": s(r[9]),
                                                       "gerencia": s(r[2]), "contrato": s(r[11])}
            upsert_fuentes(conn, extra, reporte_id)

            n, sk = load_fact_programa(conn, wb["BDP_Programa"], reporte_id, (vice, tipo))
            log(conn, reporte_id, "BDP_Programa", "core.fact_programa_ecp", n+sk, n)
            print(f"  fact_programa: {n:,} filas ({sk} descartadas)")

        # ---- COMENTARIOS ----
        if "COMENTARIOS" in wb.sheetnames:
            n = load_comentarios(conn, wb["COMENTARIOS"], reporte_id, tipo)
            log(conn, reporte_id, "COMENTARIOS", "core.fact_comentarios_produccion", n, n)
            print(f"  comentarios: {n:,} filas")

        # ---- FILIALES / POP / PROMEDIOS / CONFIG (de TODOS los archivos) ----
        if "Producción filiales" in wb.sheetnames:
            n = load_filiales(conn, wb["Producción filiales"], reporte_id, emp, tipo, treg)
            log(conn, reporte_id, "Producción filiales", "core.fact_produccion_diaria", n, n)
            print(f"  fact_produccion_diaria: {n:,} filas")
        if "POP Filiales y Exploración" in wb.sheetnames:
            n = load_pop(conn, wb["POP Filiales y Exploración"], reporte_id, emp)
            log(conn, reporte_id, "POP Filiales y Exploración", "core.fact_plan_mensual", n, n)
            print(f"  fact_plan_mensual: {n:,} filas")
        if "INICIO" in wb.sheetnames:
            n = load_promedios(conn, wb["INICIO"], reporte_id, emp, tipo)
            log(conn, reporte_id, "INICIO", "core.fact_promedio_validado", n, n)
            print(f"  fact_promedio_validado: {n:,} filas")
            v = update_config_inicio(conn, wb["INICIO"], reporte_id)
            print(f"  config_reporte actualizado: corte={v['fc']} mes={v['mi']}..{v['mf']} "
                  f"version={v['vs']} dias_anio={v['da']}")

    wb.close()
    print(f"\nOK en {time.time()-t0:,.1f}s")

if __name__ == "__main__":
    main()
