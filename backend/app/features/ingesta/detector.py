"""Detección raw/STD por presencia de hoja."""
import zipfile
from xml.etree import ElementTree as ET
from openpyxl import load_workbook

RAW = {"BDP_datos_dia", "BDP_datos_mes", "BDP_Programa"}

def tiene_raw(sheetnames: set[str]) -> bool:
    return RAW.issubset(sheetnames)

_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

def nombres_de_hojas(path) -> set[str]:
    """Nombres de hoja de un .xlsm leyendo SOLO xl/workbook.xml del zip (sin cargar celdas).
    Rápido incluso en archivos de 125 MB. Devuelve set vacío si el archivo no es un OOXML válido."""
    try:
        with zipfile.ZipFile(path) as z:
            xml = z.read("xl/workbook.xml")
    except (zipfile.BadZipFile, KeyError, OSError):
        return set()
    root = ET.fromstring(xml)
    return {sh.get("name") for sh in root.iter(f"{{{_NS_MAIN}}}sheet") if sh.get("name")}
