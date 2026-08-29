"""
Servicio de ingesta — lógica copiada del prototipo ETL (H1: el original no se toca).
Refactores aplicados:
  H2: usa app.core.db.get_engine(); prohibido load_dotenv/create_engine propios.
  H6: structlog en lugar de print().
  H7: sin flag --solo-derivadas; siempre ingesta completa.
  H3: acumula conteos en filas_por_tabla y devuelve ResultadoIngesta.
"""
from __future__ import annotations
import re, json
import datetime as dt
from pathlib import Path

import sqlalchemy as sa
from openpyxl import load_workbook

from app.core.config import get_settings
from app.core.db import get_engine
from app.core.logging import log
from app.features.ingesta.detector import tiene_raw
from app.features.ingesta.schemas import ResultadoIngesta
from app.features.ingesta.transforms import (
    BZ_DIA, BZ_MES, BZ_PRG,
    EMP_NORM, PROD_NORM, norm_emp, norm_prod, split_label,
)
from app.shared.utils import NOISE, s, num, to_date

CHUNK = 10_000

# ---------------------------------------------------------------- dims pequeñas (cache + upsert)
class DimCache:
    """Resuelve nombre->id de una dim con UNIQUE(nombre/codigo); upsert si falta."""
    def __init__(self, conn, table, idcol, namecol):
        self.conn, self.table, self.idcol, self.namecol = conn, table, idcol, namecol
        self.cache = {}
        for r in conn.execute(sa.text(f"SELECT {idcol}, {namecol} FROM {table}")):
            self.cache[r[1]] = r[0]

    def get(self, name, extra_cols=""):
        if name is None: return None
        if name in self.cache: return self.cache[name]
        rid = self.conn.execute(sa.text(
            f"INSERT INTO {self.table} ({self.namecol}{extra_cols}) VALUES (:n{',NULL'*extra_cols.count(',')}) "
            f"ON CONFLICT ({self.namecol}) DO UPDATE SET {self.namecol}=EXCLUDED.{self.namecol} "
            f"RETURNING {self.idcol}"), {"n": name}).scalar()
        self.cache[name] = rid
        return rid

# ---------------------------------------------------------------- dim_fuente / dim_fecha
FUENTE_COLS = ["nombre","campo","contrato","tipo_contrato","operador","modalidad","operacion",
               "nacionalidad","gerencia","grupo1","grupo2","grupo3","activos","fuente_contrato"]

def upsert_fuentes(conn, fuentes: dict, reporte_id: int):
    if not fuentes: return
    sets = ", ".join(f"{c}=COALESCE(EXCLUDED.{c}, core.dim_fuente.{c})" for c in FUENTE_COLS)
    cols = ", ".join(FUENTE_COLS)
    binds = ", ".join(f":{c}" for c in FUENTE_COLS)
    stmt = sa.text(f"""
        INSERT INTO core.dim_fuente (fuente_id, {cols}, reporte_id_origen)
        VALUES (:fuente_id, {binds}, :rep)
        ON CONFLICT (fuente_id) DO UPDATE SET {sets},
            reporte_id_origen = EXCLUDED.reporte_id_origen, updated_at = now()""")
    rows = []
    for fid, a in fuentes.items():
        d = {"fuente_id": fid, "rep": reporte_id}
        for c in FUENTE_COLS: d[c] = a.get(c)
        rows.append(d)
    conn.execute(stmt, rows)

def ensure_fechas(conn, fechas: set):
    rows = []
    for f in fechas:
        if f is None: continue
        rows.append({"f": f, "a": f.year, "m": f.month, "d": f.day,
                     "t": (f.month-1)//3+1, "dw": f.isoweekday(), "sm": (f.day-1)//7+1})
    if rows:
        conn.execute(sa.text("""
            INSERT INTO core.dim_fecha (fecha,anio,mes,dia,trimestre,dia_semana,semana_mes)
            VALUES (:f,:a,:m,:d,:t,:dw,:sm) ON CONFLICT (fecha) DO NOTHING"""), rows)

# ---------------------------------------------------------------- detección + config
def get_reporte(conn, path: Path, tiene_raw_: bool):
    m = re.search(r"(\d{8})", path.name)
    fecha_rep = to_date(m.group(1)) if m else None
    tipo = "NEW" if tiene_raw_ else "STD"
    nivel = "FULL" if tiene_raw_ else "SIN_ECP"
    rid = conn.execute(sa.text("""
        INSERT INTO core.config_reporte (fecha_reporte, archivo_nombre, tipo_archivo, tiene_raw, nivel_detalle)
        VALUES (:fr, :an, :tp, :raw, :nv)
        ON CONFLICT (fecha_reporte) DO UPDATE SET
            archivo_nombre=EXCLUDED.archivo_nombre, tipo_archivo=EXCLUDED.tipo_archivo,
            tiene_raw=EXCLUDED.tiene_raw, nivel_detalle=EXCLUDED.nivel_detalle, ingested_at=now()
        RETURNING reporte_id"""),
        {"fr": fecha_rep, "an": path.name, "tp": tipo, "raw": tiene_raw_, "nv": nivel}).scalar()
    return rid, fecha_rep

def _log_ingesta(conn, reporte_id, hoja, destino, leidas, ins, estado="OK", msg=None):
    conn.execute(sa.text("""
        INSERT INTO core.ingesta_log (reporte_id,hoja,tabla_destino,filas_leidas,filas_insertadas,estado,mensaje)
        VALUES (:r,:h,:d,:l,:i,:e,:m)"""),
        {"r": reporte_id, "h": hoja, "d": destino, "l": leidas, "i": ins, "e": estado, "m": msg})

# ---------------------------------------------------------------- BRONZE
def land_bronze_typed(conn, ws, table, cols, reporte_id):
    """Inserta una hoja plana raw en su tabla bronze tipada (todo TEXT). Pasada por chunks."""
    conn.execute(sa.text(f"DELETE FROM bronze.{table} WHERE reporte_id=:r"), {"r": reporte_id})
    binds = ", ".join(f":{c}" for c in cols)
    stmt = sa.text(f"INSERT INTO bronze.{table} (reporte_id, fila_origen, {', '.join(cols)}) "
                   f"VALUES (:_rep, :_fila, {binds})")
    buf, total = [], 0
    it = ws.iter_rows(values_only=True); next(it, None)  # salta encabezado
    for i, row in enumerate(it, start=2):
        if row is None or all(c is None for c in row): continue
        d = {"_rep": reporte_id, "_fila": i}
        for j, c in enumerate(cols): d[c] = (None if j >= len(row) or row[j] is None else str(row[j]))
        buf.append(d); total += 1
        if len(buf) >= CHUNK: conn.execute(stmt, buf); buf.clear()
    if buf: conn.execute(stmt, buf)
    return total

def land_landing(conn, ws, hoja, reporte_id):
    """Aterriza cualquier otra hoja a bronze.hoja_landing (JSONB), preservando celdas."""
    conn.execute(sa.text("DELETE FROM bronze.hoja_landing WHERE reporte_id=:r AND hoja=:h"),
                 {"r": reporte_id, "h": hoja})
    it = ws.iter_rows(values_only=True)
    header = next(it, None) or ()
    keys = [str(h) if h is not None else f"col{j}" for j, h in enumerate(header)]
    stmt = sa.text("""INSERT INTO bronze.hoja_landing (reporte_id,hoja,fila_origen,payload)
                      VALUES (:r,:h,:f, CAST(:p AS jsonb))""")
    buf, total = [], 0
    for i, row in enumerate(it, start=2):
        if row is None or all(c is None for c in row): continue
        payload = {keys[j] if j < len(keys) else f"col{j}":
                   (None if v is None else str(v)) for j, v in enumerate(row)}
        buf.append({"r": reporte_id, "h": hoja, "f": i, "p": json.dumps(payload, ensure_ascii=False)})
        total += 1
        if len(buf) >= CHUNK: conn.execute(stmt, buf); buf.clear()
    if buf: conn.execute(stmt, buf)
    return total

# ---------------------------------------------------------------- FACTS ECP
def _header_idx(ws, it):
    """Mapa {NOMBRE_COLUMNA: indice} desde la fila de encabezado (ya consumida de `it`).
    Resuelve por nombre en vez de posicion fija: el layout de BDP_datos_dia/mes ha cambiado
    de tamaño/orden entre vintages del reporte (ver hallazgo 2026-07-06, corpus Febrero 2026:
    dia paso de 30->32 cols, mes de 59->31 cols con reorden completo)."""
    return {str(h).strip().upper(): i for i, h in enumerate(next(it)) if h is not None}

def load_fact_dia(conn, ws, reporte_id, caches):
    DIA = sa.text("""
      INSERT INTO core.fact_produccion_dia_ecp
        (fecha,fuente_id,vice_id,socio_id,concepto_id,tipo_producto_id,producto,grupo_prod,
         propietario,volumen,porcentaje,voldismez,vol_estimado,promedio,reporte_id)
      VALUES (:fecha,:fuente_id,:vice_id,:socio_id,:concepto_id,:tipo_producto_id,:producto,
              :grupo_prod,:propietario,:volumen,:porcentaje,:voldismez,:vol_estimado,:promedio,:rep)
      ON CONFLICT ON CONSTRAINT uk_dia DO UPDATE SET
        volumen=EXCLUDED.volumen, porcentaje=EXCLUDED.porcentaje, voldismez=EXCLUDED.voldismez,
        vol_estimado=EXCLUDED.vol_estimado, promedio=EXCLUDED.promedio, reporte_id=EXCLUDED.reporte_id""")
    vice, socio, conc, tipo = caches
    it = ws.iter_rows(values_only=True)
    idx = _header_idx(ws, it)
    i_idbdp, i_fecha = idx["IDBDP"], idx["FECHA"]
    i_fuente, i_contrato, i_tipocontrato = idx["FUENTE"], idx["CONTRATO"], idx["TIPOCONTRATO"]
    i_operador, i_modalidad, i_operacion = idx["OPERADOR"], idx["MODALIDAD"], idx["OPERACION"]
    i_nacionalidad, i_gerencia = idx["NACIONALIDAD"], idx["GERENCIA"]
    i_grupo1, i_grupo2, i_grupo3 = idx["GRUPO1"], idx["GRUPO2"], idx["GRUPO3"]
    i_activos, i_fuentecontrato = idx["ACTIVOS"], idx["FUENTECONTRATO"]
    i_vice, i_socio, i_concepto = idx["GRUPO1_SIGLA"], idx["SOCIO"], idx["CONCEPTO"]
    i_tipoprod, i_producto = idx["TIPOPRODUCTO"], idx["PRODUCTO"]
    i_grupoprod, i_propietario = idx["GRUPOPROD"], idx["PROPIETARIO"]
    i_volumen, i_porcentaje = idx["VOLUMEN"], idx["PORCENTAJE"]
    i_voldismez, i_vol_estimado, i_promedio = idx["VOLDISMEZ"], idx["VOL_ESTIMADO"], idx["PROMEDIO"]

    buf, fuentes, fechas, total, skip = [], {}, set(), 0, 0
    for row in it:
        if row is None or row[i_idbdp] is None: continue
        fecha = to_date(row[i_fecha]); idbdp = num(row[i_idbdp])
        if fecha is None or idbdp is None: skip += 1; continue
        fid = int(idbdp)
        fuentes[fid] = {"nombre": s(row[i_fuente]), "contrato": s(row[i_contrato]),
                        "tipo_contrato": s(row[i_tipocontrato]), "operador": s(row[i_operador]),
                        "modalidad": s(row[i_modalidad]), "operacion": s(row[i_operacion]),
                        "nacionalidad": s(row[i_nacionalidad]), "gerencia": s(row[i_gerencia]),
                        "grupo1": s(row[i_grupo1]), "grupo2": s(row[i_grupo2]), "grupo3": s(row[i_grupo3]),
                        "activos": s(row[i_activos]), "fuente_contrato": s(row[i_fuentecontrato])}
        fechas.add(fecha)
        buf.append({"fecha": fecha, "fuente_id": fid, "vice_id": vice.get(s(row[i_vice])),
                    "socio_id": socio.get(s(row[i_socio])), "concepto_id": conc.get(s(row[i_concepto])),
                    "tipo_producto_id": tipo.get(s(row[i_tipoprod])), "producto": s(row[i_producto]) or "",
                    "grupo_prod": s(row[i_grupoprod]) or "", "propietario": s(row[i_propietario]) or "",
                    "volumen": num(row[i_volumen]), "porcentaje": num(row[i_porcentaje]),
                    "voldismez": num(row[i_voldismez]), "vol_estimado": num(row[i_vol_estimado]),
                    "promedio": num(row[i_promedio]), "rep": reporte_id})
        total += 1
        if len(buf) >= CHUNK:
            ensure_fechas(conn, fechas); upsert_fuentes(conn, fuentes, reporte_id)
            conn.execute(DIA, buf); buf.clear(); fuentes.clear(); fechas.clear()
    if buf:
        ensure_fechas(conn, fechas); upsert_fuentes(conn, fuentes, reporte_id); conn.execute(DIA, buf)
    return total, skip

def load_fact_mes(conn, ws, reporte_id, caches, emit=None):
    MES = sa.text("""
      INSERT INTO core.fact_produccion_mes_ecp
        (fecha,fuente_id,vice_id,socio_id,concepto_id,tipo_producto_id,producto,escenario_id,
         proceso_id,grupo_prod,negocio,volumen,porcentaje,voldismez,bpd_m,bpda_ac,bpd_a,
         bpdeq_m,blseq,bpdeq_a,reporte_id)
      VALUES (:fecha,:fuente_id,:vice_id,:socio_id,:concepto_id,:tipo_producto_id,:producto,
              :escenario_id,:proceso_id,:grupo_prod,:negocio,:volumen,:porcentaje,:voldismez,
              :bpd_m,:bpda_ac,:bpd_a,:bpdeq_m,:blseq,:bpdeq_a,:rep)
      ON CONFLICT ON CONSTRAINT uk_mes DO UPDATE SET
        volumen=EXCLUDED.volumen, porcentaje=EXCLUDED.porcentaje, voldismez=EXCLUDED.voldismez,
        bpd_m=EXCLUDED.bpd_m, bpda_ac=EXCLUDED.bpda_ac, bpd_a=EXCLUDED.bpd_a, bpdeq_m=EXCLUDED.bpdeq_m,
        blseq=EXCLUDED.blseq, bpdeq_a=EXCLUDED.bpdeq_a, negocio=EXCLUDED.negocio, reporte_id=EXCLUDED.reporte_id""")
    vice, socio, conc, tipo, esc, proc = caches
    it = ws.iter_rows(values_only=True)
    idx = _header_idx(ws, it)
    opt = idx.get  # columnas que pueden faltar en este vintage de la hoja -> None sin reventar
    i_idbdp, i_fecha = idx["IDBDP"], idx["FECHA"]
    i_fuente, i_contrato, i_operador = opt("FUENTE"), opt("CONTRATO"), opt("OPERADOR")
    i_gerencia, i_grupo1, i_grupo2 = opt("GERENCIA"), opt("GRUPO1"), opt("GRUPO2")
    i_activos, i_fuentecontrato = opt("ACTIVOS"), opt("FUENTECONTRATO")
    i_vice, i_socio, i_concepto = opt("NIVEL1_SIGLA"), idx["SOCIO"], idx["CONCEPTO"]
    i_tipoprod, i_producto = idx["TIPOPRODUCTO"], opt("PRODUCTO")
    i_escenario, i_proceso = idx["ESCENARIO"], idx["PROCESO"]
    i_grupoprod, i_negocio = idx["GRUPOPROD"], opt("NEGOCIO")
    i_volumen, i_porcentaje, i_voldismez = idx["VOLUMEN"], idx["PORCENTAJE"], idx["VOLDISMEZ"]
    i_bpd_m, i_bpda_ac, i_bpd_a = idx["BPD_M"], idx["BPDA_AC"], idx["BPD_A"]
    i_bpdeq_m, i_blseq, i_bpdeq_a = idx["BPDEQ_M"], idx["BLSEQ"], idx["BPDEQ_A"]

    def g(i, row):
        return row[i] if i is not None else None

    buf, fuentes, fechas, total, skip = [], {}, set(), 0, 0
    for row in it:
        if row is None or row[i_idbdp] is None: continue
        fecha = to_date(row[i_fecha]); idbdp = num(row[i_idbdp])
        if fecha is None or idbdp is None: skip += 1; continue
        fid = int(idbdp)
        fuentes[fid] = {"nombre": s(g(i_fuente, row)), "contrato": s(g(i_contrato, row)),
                        "operador": s(g(i_operador, row)), "gerencia": s(g(i_gerencia, row)),
                        "grupo1": s(g(i_grupo1, row)), "grupo2": s(g(i_grupo2, row)),
                        "activos": s(g(i_activos, row)), "fuente_contrato": s(g(i_fuentecontrato, row))}
        fechas.add(fecha)
        buf.append({"fecha": fecha, "fuente_id": fid, "vice_id": vice.get(s(g(i_vice, row))),
                    "socio_id": socio.get(s(row[i_socio])), "concepto_id": conc.get(s(row[i_concepto])),
                    "tipo_producto_id": tipo.get(s(row[i_tipoprod])), "producto": s(g(i_producto, row)) or "",
                    "escenario_id": esc.get(s(row[i_escenario])), "proceso_id": proc.get(s(row[i_proceso])),
                    "grupo_prod": s(row[i_grupoprod]) or "", "negocio": s(g(i_negocio, row)),
                    "volumen": num(row[i_volumen]), "porcentaje": num(row[i_porcentaje]), "voldismez": num(row[i_voldismez]),
                    "bpd_m": num(row[i_bpd_m]), "bpda_ac": num(row[i_bpda_ac]), "bpd_a": num(row[i_bpd_a]),
                    "bpdeq_m": num(row[i_bpdeq_m]), "blseq": num(row[i_blseq]), "bpdeq_a": num(row[i_bpdeq_a]), "rep": reporte_id})
        total += 1
        if len(buf) >= CHUNK:
            ensure_fechas(conn, fechas); upsert_fuentes(conn, fuentes, reporte_id)
            conn.execute(MES, buf); buf.clear(); fuentes.clear(); fechas.clear()
            log.info("ingesta.mes.chunk", filas=total)
            if emit: emit({"tipo": "avance", "hoja": "BDP_datos_mes",
                           "tabla": "fact_produccion_mes_ecp", "filas": total})
    if buf:
        ensure_fechas(conn, fechas); upsert_fuentes(conn, fuentes, reporte_id); conn.execute(MES, buf)
    return total, skip

def load_fact_programa(conn, ws, reporte_id, caches):
    PRG = sa.text("""
      INSERT INTO core.fact_programa_ecp
        (fecha,vice_id,tipo_producto_id,fuente_id,area,campo,version,fecha_version,
         volumen,produccion_total,part_ecp,reporte_id)
      VALUES (:fecha,:vice_id,:tipo_producto_id,:fuente_id,:area,:campo,:version,:fecha_version,
              :volumen,:produccion_total,:part_ecp,:rep)
      ON CONFLICT ON CONSTRAINT uk_prog DO UPDATE SET
        volumen=EXCLUDED.volumen, produccion_total=EXCLUDED.produccion_total,
        part_ecp=EXCLUDED.part_ecp, fecha_version=EXCLUDED.fecha_version, reporte_id=EXCLUDED.reporte_id""")
    vice, tipo = caches
    buf, fechas, total, skip = [], set(), 0, 0
    it = ws.iter_rows(values_only=True); next(it, None)
    for row in it:
        if row is None or row[0] is None: continue
        fecha = to_date(row[0])
        if fecha is None: skip += 1; continue
        idbdp = num(row[10]); fid = int(idbdp) if idbdp is not None else None
        fechas.add(fecha)
        buf.append({"fecha": fecha, "vice_id": vice.get(s(row[1])),
                    "tipo_producto_id": tipo.get(s(row[8])), "fuente_id": fid,
                    "area": s(row[9]) or "", "campo": s(row[7]) or "", "version": s(row[3]) or "",
                    "fecha_version": to_date(row[4]), "volumen": num(row[6]),
                    "produccion_total": num(row[12]), "part_ecp": num(row[13]), "rep": reporte_id})
        total += 1
        if len(buf) >= CHUNK:
            ensure_fechas(conn, fechas); conn.execute(PRG, buf); buf.clear(); fechas.clear()
    if buf:
        ensure_fechas(conn, fechas); conn.execute(PRG, buf)
    return total, skip

# ---------------------------------------------------------------- COMENTARIOS
def _coment_cell(row, i):
    """Celda de comentario en posición i: limpia ruido (#REF!/#N/A/(en blanco)) y trata '0' como vacío."""
    v = s(row[i]) if len(row) > i else None
    return None if v == "0" else v


def load_comentarios(conn, ws, reporte_id, tipo_cache):
    """'COMENTARIOS' -> core.fact_comentarios_produccion. 1 tabla continua (CRUDO/GAS/BLANCOS x activos x
    area) con 3 campos de texto: D=comentario (real), E=comentario_programa (COMENTARIO PROGRAMA, header f1),
    G=comentario_extra (disperso). Forward-fill de PRODUCTO (col A dispersa, p.ej. fila con A en blanco).
    'comentario' (NOT NULL) usa la cadena de respaldo D->E->G. Trata '0' y ruido como vacío (corrige el bug
    de cortocircuito '0' del modelo previo). Idempotente: DELETE por reporte_id + INSERT (sin UNIQUE)."""
    conn.execute(sa.text("DELETE FROM core.fact_comentarios_produccion WHERE reporte_id=:r"), {"r": reporte_id})
    stmt = sa.text("""INSERT INTO core.fact_comentarios_produccion
                      (tipo_producto_id, activos, area, comentario, comentario_programa, comentario_extra, reporte_id)
                      VALUES (:t, :a, :ar, :c, :cp, :ce, :r)""")
    buf, total, producto = [], 0, None
    it = ws.iter_rows(values_only=True); next(it, None)        # salta el encabezado (fila 1)
    for row in it:
        if row is None or len(row) < 4:
            continue
        a = s(row[0])
        if a:
            producto = a                                       # forward-fill del PRODUCTO (col A dispersa)
        com   = _coment_cell(row, 3)                           # D  comentario real
        prog  = _coment_cell(row, 4)                           # E  COMENTARIO PROGRAMA
        extra = _coment_cell(row, 6)                           # G  comentario extra (disperso)
        principal = com or prog or extra                       # comentario es NOT NULL -> respaldo D->E->G
        if not principal:
            continue
        buf.append({"t": tipo_cache.get(producto), "a": s(row[1]), "ar": s(row[2]),
                    "c": principal, "cp": prog, "ce": extra, "r": reporte_id})
        total += 1
    if buf:
        conn.execute(stmt, buf)
    return total

# ---------------------------------------------------------------- FILIALES / POP / PROMEDIOS / CONFIG
def load_filiales(conn, ws, reporte_id, emp, prod, treg):
    """Bloques REAL y PROGRAMA de 'Producción filiales' → fact_produccion_diaria (unpivot)."""
    stmt = sa.text("""
      INSERT INTO core.fact_produccion_diaria (empresa_id,producto_id,tipo_id,fecha,valor_produccion,reporte_id)
      VALUES (:e,:p,:t,:f,:v,:r)
      ON CONFLICT ON CONSTRAINT uk_fil DO UPDATE SET
        valor_produccion=EXCLUDED.valor_produccion, reporte_id=EXCLUDED.reporte_id""")
    tipo, dates, buf, fechas, total = None, [], [], set(), 0
    for row in ws.iter_rows(values_only=True):
        a = s(row[0]) if row else None
        if a is None: continue
        au = a.upper()
        if au == "REAL": tipo, dates = "Real", []; continue
        if au == "PROGRAMA": tipo, dates = "Programa", []; continue
        if au.startswith("PROYEC"): tipo, dates = None, []; continue
        if au == "EMPRESA": dates = [to_date(v) for v in row[1:]]; continue
        if tipo is None or au.startswith("TOTAL"): continue
        emp_n, prod_n = split_label(a)
        eid, pid, tid = emp.get(norm_emp(emp_n)), prod.get(norm_prod(prod_n)), treg.get(tipo)
        if not (eid and pid and tid): continue
        for j, val in enumerate(row[1:]):
            f = dates[j] if j < len(dates) else None
            v = num(val)
            if f is None or v is None: continue
            fechas.add(f)
            buf.append({"e": eid, "p": pid, "t": tid, "f": f, "v": v, "r": reporte_id})
    ensure_fechas(conn, fechas)
    if buf: conn.execute(stmt, buf)
    return len(buf)

def load_pop(conn, ws, reporte_id, emp):
    """'POP Filiales y Exploración' filas TOTAL <empresa> → fact_plan_mensual (pop_kbd)."""
    stmt = sa.text("""
      INSERT INTO core.fact_plan_mensual (empresa_id,anio,mes,pop_kbd,segmento,reporte_id)
      VALUES (:e,:a,:m,:p,'Filiales',:r)
      ON CONFLICT ON CONSTRAINT uk_plan DO UPDATE SET pop_kbd=EXCLUDED.pop_kbd, reporte_id=EXCLUDED.reporte_id""")
    dates, buf = [], []
    for row in ws.iter_rows(values_only=True):
        if row is None: continue
        if s(row[1]) == "Producto":                       # fila de encabezado con fechas
            dates = [to_date(v) for v in row]
            continue
        b = s(row[1])
        emp_n = norm_emp(s(row[2]))
        if not (b and b.upper().startswith("TOTAL") and emp_n in ("Hocol", "America", "Permian")):
            continue
        eid = emp.get(emp_n)
        for j, val in enumerate(row):
            f = dates[j] if j < len(dates) else None
            v = num(val)
            if f is None or v is None: continue
            buf.append({"e": eid, "a": f.year, "m": f.month, "p": v/1000.0, "r": reporte_id})
    if buf: conn.execute(stmt, buf)
    return len(buf)

def load_promedios(conn, ws, reporte_id, emp, prod):
    """Sección 'REAL PROMEDIO MES (YTD)' de INICIO → fact_promedio_validado."""
    stmt = sa.text("""
      INSERT INTO core.fact_promedio_validado (empresa_id,producto_id,anio,mes,promedio_validado,reporte_id)
      VALUES (:e,:p,:a,:m,:v,:r)
      ON CONFLICT ON CONSTRAINT uk_prom DO UPDATE SET
        promedio_validado=EXCLUDED.promedio_validado, reporte_id=EXCLUDED.reporte_id""")
    dates, buf, in_sec = [], [], False
    for row in ws.iter_rows(values_only=True):
        a = s(row[0]) if row else None
        if a == "Producto" and s(row[1]) == "Empresa":
            dates = [to_date(v) for v in row]; in_sec = True; continue
        if not in_sec: continue
        if a is None or a.upper() == "TOTAL": continue
        pid = prod.get(norm_prod(a)); eid = emp.get(norm_emp(s(row[1])))
        if not (pid and eid): continue
        for j, val in enumerate(row):
            f = dates[j] if j < len(dates) else None
            v = num(val)
            if f is None or v is None: continue
            buf.append({"e": eid, "p": pid, "a": f.year, "m": f.month, "v": v, "r": reporte_id})
    if buf: conn.execute(stmt, buf)
    return len(buf)

def update_config_inicio(conn, ws, reporte_id):
    """Completa fecha_corte/mes_inicio/mes_fin/version_semana/anio_inicio/dias_anio desde INICIO."""
    lab = {}
    for row in ws.iter_rows(min_row=1, max_row=40, max_col=4, values_only=True):
        b = s(row[1])
        if b: lab[b] = row[2]
    def n(x): v = num(x); return int(v) if v is not None else None
    vals = {
        "fc": to_date(lab.get("Día de corte")),
        "mi": to_date(lab.get("1er dia del mes")),
        "mf": to_date(lab.get("MES CORTE")),
        "vs": n(lab.get("Version Semana")),
        "ai": (to_date(lab.get("Fecha inicial año")).year if to_date(lab.get("Fecha inicial año")) else None),
        "da": n(lab.get("Días del año")),
        "r": reporte_id}
    conn.execute(sa.text("""
        UPDATE core.config_reporte SET fecha_corte=:fc, mes_inicio=:mi, mes_fin=:mf,
            version_semana=:vs, anio_inicio=:ai, dias_anio=:da WHERE reporte_id=:r"""), vals)
    return vals

# ---------------------------------------------------------------- helper batch (H8)
import re as _re

def _archivos_ordenados(data_dir: Path) -> list[Path]:
    """Todos los .xlsm bajo data_dir ordenados por la fecha YYYYMMDD del nombre (asc)."""
    def fecha(p: Path) -> str:
        m = _re.search(r"(\d{8})", p.name); return m.group(1) if m else "00000000"
    return sorted(data_dir.rglob("*.xlsm"), key=fecha)

# ---------------------------------------------------------------- punto de entrada público (H3)
RAW_SHEETS = {"BDP_datos_dia", "BDP_datos_mes", "BDP_Programa"}

def _p50_contig_months(grid, hdr_row, start_col):
    """Columnas de mes CONTIGUAS desde start_col en hdr_row; corta en la 1ª no-fecha.
    Crítico (auditoría A4): evita cruzar a la tabla VR/GER que comparte fila de encabezado.
    Usa to_date (helper del repo): int YYYYMMDD / datetime / '20240930.0' -> date; otro -> None."""
    out, c = [], start_col
    while True:
        d = to_date(grid.get((hdr_row, c)))
        if d is None:
            break
        out.append((c, d))
        c += 1
    return out

def _p50_grid(ws):
    grid, maxr = {}, 0
    for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for c, v in enumerate(row, start=1):
            if v is not None and str(v).strip() != "":
                grid[(r, c)] = v
                if r > maxr:
                    maxr = r
        if r > 250:
            break
    return grid, maxr

def _p50_extract(ws):
    """Extractor de la hoja 'P50 Quemado <año> ECP y Filiales'. Contrato genérico:
    devuelve [{tabla_idx, tabla_label, dims(dict), fecha(date), valor(float)}].
    Tabla 1='quemado' (escenario/producto/vice/activos/area); Tabla 2='filiales' (producto/empresa).
    Lee por posición; meses contiguos (corta antes de la tabla VR/GER); descarta subtotales y Promedio Año."""
    grid, maxr = _p50_grid(ws)
    rows = []
    # --- Tabla 1 ---
    m1 = _p50_contig_months(grid, 2, 6)
    for r in range(3, maxr + 1):
        area = grid.get((r, 5))
        if area is None or str(area).strip().lower().startswith("total"):
            continue
        dims = {"escenario": s(grid.get((r, 1))), "producto": s(grid.get((r, 2))),
                "vice": s(grid.get((r, 3))), "activos": s(grid.get((r, 4))), "area": s(area)}
        for c, d in m1:
            val = num(grid.get((r, c)))
            if val is None:
                continue
            rows.append({"tabla_idx": 1, "tabla_label": "Tabla 1", "dims": dims, "fecha": d, "valor": val})
    # --- Tabla 2 ---
    title = next(((r, c) for (r, c), v in grid.items()
                  if isinstance(v, str) and v.strip().lower() == "p50 filiales"), None)
    if title:
        tr, tc = title
        hdr = tr + 1
        m2 = _p50_contig_months(grid, hdr, tc + 3)
        for r in range(hdr + 1, maxr + 1):
            empresa = grid.get((r, tc + 2))
            if empresa is None:
                continue
            producto = grid.get((r, tc))
            if producto and str(producto).strip().lower().startswith("total"):
                continue
            dims = {"producto": s(producto), "empresa": s(empresa)}
            for c, d in m2:
                val = num(grid.get((r, c)))
                if val is None:
                    continue
                rows.append({"tabla_idx": 2, "tabla_label": "Tabla 2", "dims": dims, "fecha": d, "valor": val})
    return rows

def _filiales_extract(ws):
    """Extractor de 'Producción filiales' → 8 tablas. Contrato genérico [{tabla_idx,tabla_label,dims,fecha,valor}].
    Familia A — columnas = FECHAS (reusa split_label/norm_emp/norm_prod/to_date):
      1 REAL, 2 PROGRAMA, 3 PROYECCIÓN (empresa×producto); 6 REAL, 7 PROGRAMA (totales por empresa,
      sin producto). La tabla 7 trae el header de fechas vacío → reusa las fechas de la 6.
    Familia B — columnas = CATEGORÍAS (matriz, fecha=NULL, dims={fila,columna}):
      4 FILIALES mes/semana, 5 Seguimiento semanal, 8 Desempeño P50.
    Layout verificado estable (anclajes idénticos en 4 archivos de muestra)."""
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    n = len(grid)
    rows = []

    def cell(i, j):
        return grid[i][j] if 0 <= i < n and 0 <= j < len(grid[i]) else None
    def clean(x):
        t = s(x)
        return re.sub(r"\s+", " ", t) if t else ""
    def ffill(seq):
        out, last = [], ""
        for v in seq:
            c = clean(v)
            if c:
                last = c
            out.append(last)
        return out

    IDX_PROD = {"REAL": (1, "Tabla 1 (REAL)"), "PROGRAMA": (2, "Tabla 2 (PROGRAMA)"),
                "PROYECC": (3, "Tabla 3 (PROYECCIÓN)")}
    IDX_EMP = {"REAL": (6, "Tabla 6 (REAL total empresa)"),
               "PROGRAMA": (7, "Tabla 7 (PROGRAMA total empresa)")}
    EMPRESAS = {"HOCOL", "AMERICA", "PERMIAN", "EAI", "EA"}

    # ---------- Familia A: bloques diarios (encabezado 'EMPRESA' con fechas) ----------
    last_emp_dates = None
    i = 0
    while i < n:
        au = clean(cell(i, 0)).upper()
        base = "REAL" if au == "REAL" else ("PROGRAMA" if au == "PROGRAMA"
               else ("PROYECC" if au.startswith("PROYECC") else None))
        if base is None:
            i += 1; continue
        j = i + 1                                   # el header de fechas ('EMPRESA') va debajo
        while j < n and clean(cell(j, 0)) == "":
            j += 1
        if clean(cell(j, 0)).upper() != "EMPRESA":
            i += 1; continue                        # no es bloque diario (p.ej. 'FILIALES')
        dates = [to_date(v) for v in grid[j][1:]]
        if not any(d is not None for d in dates):
            dates = None                            # header de fechas vacío (Tabla 7) → reusar
        k = j + 1
        while k < n:
            la = clean(cell(k, 0)); lau = la.upper()
            if la == "" or lau == "EMPRESA" or lau in ("REAL", "PROGRAMA") or lau.startswith("PROYECC"):
                break
            if lau.startswith("TOTAL"):
                k += 1; continue
            emp_raw, prod_raw = split_label(la)
            if emp_raw and prod_raw:                # nivel producto → tablas 1/2/3
                empresa, producto = norm_emp(emp_raw), norm_prod(prod_raw)
                if empresa and producto and dates:
                    idx, label = IDX_PROD[base]
                    for jj, val in enumerate(grid[k][1:]):
                        f = dates[jj] if jj < len(dates) else None
                        v = num(val)
                        if f is None or v is None:
                            continue
                        rows.append({"tabla_idx": idx, "tabla_label": label,
                                     "dims": {"empresa": empresa, "producto": producto},
                                     "fecha": f, "valor": v})
            elif lau in EMPRESAS and base in IDX_EMP:   # nivel empresa (total) → tablas 6/7
                empresa = norm_emp(la)
                d = dates if dates else last_emp_dates
                if empresa and d:
                    if dates:
                        last_emp_dates = dates
                    idx, label = IDX_EMP[base]
                    for jj, val in enumerate(grid[k][1:]):
                        f = d[jj] if jj < len(d) else None
                        v = num(val)
                        if f is None or v is None:
                            continue
                        rows.append({"tabla_idx": idx, "tabla_label": label,
                                     "dims": {"empresa": empresa},
                                     "fecha": f, "valor": v})
            k += 1
        i = k

    # ---------- Familia B: matrices (fecha=NULL, dims={fila,columna}) ----------
    def emit_matrix(idx, label, label_col, metric_row, period_seq, data_start):
        metrics = [clean(v) for v in grid[metric_row]] if 0 <= metric_row < n else []
        collab = {}
        for jc, m in enumerate(metrics):
            if jc == label_col or not m:
                continue
            p = period_seq[jc] if period_seq and jc < len(period_seq) else ""
            collab[jc] = (p + " " + m).strip() if p else m
        r = data_start
        while r < n:
            fila = clean(cell(r, label_col)); fu = fila.upper()
            if fu not in EMPRESAS and fu != "TOTAL":
                break
            for jc, cl in collab.items():
                v = num(cell(r, jc))
                if v is None:
                    continue
                rows.append({"tabla_idx": idx, "tabla_label": label,
                             "dims": {"fila": fila, "columna": cl}, "fecha": None, "valor": v})
            if fu == "TOTAL":                       # TOTAL cierra el bloque de la matriz
                break
            r += 1

    for r in range(n):
        a = clean(cell(r, 0)).upper()
        b = clean(cell(r, 1)).upper()
        rowtxt = " ".join(clean(c).upper() for c in grid[r])
        if a == "FILIALES" and "MES" in b:                          # #4 FILIALES mes/semana
            emit_matrix(4, "Tabla 4 (FILIALES mes/semana)", 0, r + 2, ffill(grid[r + 1]), r + 3)
        if "SEGUIMIENTO" in b:                                       # #5 Seguimiento semanal
            per = ffill([v if ("AL " in clean(v).upper() or clean(v)[:1].isdigit()) else ""
                         for v in grid[r]])
            emit_matrix(5, "Tabla 5 (Seguimiento semanal)", 0, r + 1, per, r + 2)
        if "DESEMPE" in rowtxt:                                      # #8 Desempeño P50
            emit_matrix(8, "Tabla 8 (Desempeño P50)", 3, r + 1, None, r + 2)
    return rows


def _bitacora_extract(ws):
    """Extractor de la hoja '(Bitacora)' → 3 tablas TIPOPRODUCTO×VICE × fecha (Familia A, columnas=fechas).
    DECLARA siempre las 3 tablas (1 REAL, 2 PROGRAMA, 3 PROYECCIÓN) para que el front muestre 3 ítems
    aunque PROGRAMA venga #N/A (archivos STD → 0 filas). Reusa s/num/to_date. dims={tipoproducto,vice}.
    Excluye subtotales (filas con VICE vacío: 'Total CRUDO/GAS/BLANCOS', 'Total general') y la columna
    agregada mensual a la derecha de las fechas ('REAL' → to_date None). Layout estable en 3 archivos."""
    PRODUCTOS = {"CRUDO", "GAS", "BLANCOS"}
    VICES = {"VRC", "VRO", "VAO", "VFS", "VPI", "VEX"}
    DECLARED = [(1, "Tabla 1 (REAL)"), (2, "Tabla 2 (PROGRAMA)"), (3, "Tabla 3 (PROYECCIÓN)")]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    n = len(grid)
    rows = []

    def S(i, j):
        v = grid[i][j] if 0 <= i < n and 0 <= j < len(grid[i]) else None
        return s(v) or ""

    def block_of(label):
        u = label.upper()
        if "REAL" in u: return (1, "Tabla 1 (REAL)")
        if "PROGRAMA" in u: return (2, "Tabla 2 (PROGRAMA)")
        if "PROYEC" in u: return (3, "Tabla 3 (PROYECCIÓN)")
        return None

    i = 0
    while i < n:
        a = S(i, 0)
        if not a.startswith("***"):
            i += 1; continue
        blk = block_of(a)
        if blk is None:
            i += 1; continue
        idx, label = blk
        # encabezado de fechas: la fila cuyo col A == 'TIPOPRODUCTO'
        j = i + 1
        while j < n and S(j, 0).upper() != "TIPOPRODUCTO" and not S(j, 0).startswith("***"):
            j += 1
        if j >= n or S(j, 0).upper() != "TIPOPRODUCTO":
            i += 1; continue
        dates = [to_date(v) for v in grid[j][2:]]      # desde col C; la col 'REAL' (agregado) → None
        producto = None
        k = j + 1
        while k < n and not S(k, 0).startswith("***"):
            ca = S(k, 0).upper()
            if ca in PRODUCTOS:
                producto = ca                          # forward-fill del producto (col A dispersa)
            vice = S(k, 1).upper()
            if vice in VICES and producto:             # fila de datos sii hay VICE (excluye subtotales)
                for jj, val in enumerate(grid[k][2:]):
                    f = dates[jj] if jj < len(dates) else None
                    v = num(val)
                    if f is None or v is None:         # fecha inválida / #N/A / blank → se descarta
                        continue
                    rows.append({"tabla_idx": idx, "tabla_label": label,
                                 "dims": {"tipoproducto": producto, "vice": vice},
                                 "fecha": f, "valor": v})
            k += 1
        i = k
    return {"rows": rows, "tablas": DECLARED}


def _p50_acum_extract(ws):
    """Extractor de 'P50 Acumulado' → 4 tablas de promedios ACUMULADOS ya calculados en la hoja.
    NO recalcula: toma los valores cacheados TAL COMO APARECEN (decisión usuario 2026-06-29). La hoja
    es un cálculo derivado de 'NEW MES-AÑO', pero por requerimiento se ingieren sus valores tal cual.
    Contrato genérico [{tabla_idx,tabla_label,dims,fecha,valor}]; dims={producto}.

    La hoja tiene DOS secciones (verificado en 5 archivos ene–may 2026; anclas de col A estables):
      Sección 1 'P50' base:
        Tabla 1 (P50 ECP)        título 'P50'          (r9)  → CRUDO/GAS/BLANCOS/VEX CRUDO/Total VDP
        Tabla 2 (P50 FILIALES)   título 'P50 FILIALES' (r16) → CRUDO/GAS/BLANCOS/Total Filiales
      Sección 2 'RETO CORPORATIVO' (r26 · el «compromiso», decisión usuario 2026-07-26):
        Tabla 3 (RETO CORP ECP)      título 'P50'         (r33) → mismos productos ECP
        Tabla 4 (RETO CORP FILIALES) título 'P50 FILIALES'(r40) → mismos productos filiales

    Anclado por los TÍTULOS en col A y ACOTADO por el título siguiente (next_title_after) → cada tabla se
    corta antes de la próxima ancla. 🔑 Esto arregla el bug previo (find_title+min): la Tabla 2 arrastraba
    el bloque RETO-ECP + basura (',' y 'Días del año') y conflaba Filiales con RETO-ECP bajo el mismo
    producto → colisión (dims,fecha) last-wins en la BD. La sección RETO se localiza por el 1.er título
    que empieza en 'RETO' (col A). Meses = encabezado contiguo desde col C (to_date corta en la 1.ª
    no-fecha → excluye la columna 'PROMEDIO ACUMULADO' = mes de corte /1000). Incluye las filas de total
    ('Total VDP'/'Total Filiales') tal como aparecen.

    OMITIDO (cobertura entrada 5 → salida 4, avisado al usuario 2026-07-26): el sub-bloque 'RETO 761K'
    (r48+, presente solo en 1 de 5 archivos, SIN etiquetas de producto — solo códigos VICE + totales);
    NO es el compromiso pedido. Se descarta por el acotamiento (queda tras el último título capturado).
    DECLARA siempre las 4 tablas para que el front las liste aunque vengan vacías."""
    grid, maxr = _p50_grid(ws)
    DECLARED = [(1, "Tabla 1 (P50 ECP)"), (2, "Tabla 2 (P50 FILIALES)"),
                (3, "Tabla 3 (RETO CORP ECP)"), (4, "Tabla 4 (RETO CORP FILIALES)")]
    rows = []

    def rows_titled(target):
        return sorted(r for (r, c), v in grid.items()
                      if c == 1 and isinstance(v, str) and v.strip().upper() == target)

    p50s = rows_titled("P50")                       # [9, 33] (o [9] si falta la sección RETO)
    filials = rows_titled("P50 FILIALES")           # [16, 40]
    retos = sorted(r for (r, c), v in grid.items()  # 'RETO CORPORATIVO', 'RETO 761K', …
                   if c == 1 and isinstance(v, str) and v.strip().upper().startswith("RETO"))
    rc = min(retos) if retos else maxr + 1          # fila de inicio de la sección RETO CORPORATIVO

    all_titles = sorted(set(p50s) | set(filials) | set(retos))

    def next_title_after(r):
        return next((t for t in all_titles if t > r), maxr + 1)

    def emit(idx, label, title_row):
        if title_row is None:
            return
        hdr = title_row + 1
        months = _p50_contig_months(grid, hdr, 3)      # meses contiguos desde col C
        if not months:
            return
        end = next_title_after(hdr)                    # se detiene ANTES del siguiente título
        for r in range(hdr + 1, end):
            ps = s(grid.get((r, 1)))
            if not ps:
                continue
            dims = {"producto": ps}
            for c, d in months:
                v = num(grid.get((r, c)))
                if v is None:
                    continue
                rows.append({"tabla_idx": idx, "tabla_label": label,
                             "dims": dims, "fecha": d, "valor": v})

    emit(1, "Tabla 1 (P50 ECP)",            next((r for r in p50s if r < rc), None))
    emit(2, "Tabla 2 (P50 FILIALES)",       next((r for r in filials if r < rc), None))
    emit(3, "Tabla 3 (RETO CORP ECP)",      next((r for r in p50s if r > rc), None))
    emit(4, "Tabla 4 (RETO CORP FILIALES)", next((r for r in filials if r > rc), None))
    return {"rows": rows, "tablas": DECLARED}


def _programa_extract(ws):
    """4 tablas de la hoja 'PROGRAMA' — valores cacheados TAL COMO APARECEN (NO se recalcula desde
    BDP_Programa; decisión usuario 2026-06-30). Contrato genérico [{tabla_idx,tabla_label,dims,fecha,valor}].
      T1 'Suma de Produccion_total' (pivot A:BN):   dims {tipoproducto, area, campo}
      T2 'Suma de Volumen' (pivot CI:ES):           dims {producto, vice}
      T3 'PRODUCTO/VICE' (no-pivot CI:ES, fila 31+): dims {tipoproducto, vice}
      T4 'verificador carga (detallado)' (FW:II):   dims {producto, vice, area, campo}
    Incluye las filas de total (no se filtran). Las columnas no-fecha ('(en blanco)', 'Total general')
    se omiten porque no aportan fecha. Las dims se rellenan hacia abajo respetando subtotales.
    DECLARA siempre las 4 tablas (idx 1..4) aunque vengan vacías. Anclas verificadas en el archivo NEW."""
    grid = {}
    for r, row in enumerate(ws.iter_rows(values_only=True), 1):
        for c, v in enumerate(row, 1):
            if v is not None and str(v).strip() != "":
                grid[(r, c)] = v
        if r > 330:
            break

    DECLARED = [(1, "Tabla 1 (Produccion_total)"), (2, "Tabla 2 (Volumen)"),
                (3, "Tabla 3 (Programa por VICE)"), (4, "Tabla 4 (verificador carga)")]
    # (idx, label, dims[(col, nombre)], fila_encabezado_fechas, dato_ini, dato_fin)
    TABLES = [
        (1, "Tabla 1 (Produccion_total)", [(1, "tipoproducto"), (2, "area"), (3, "campo")], 5, 6, 312),
        (2, "Tabla 2 (Volumen)", [(87, "producto"), (88, "vice")], 5, 6, 24),
        (3, "Tabla 3 (Programa por VICE)", [(87, "tipoproducto"), (88, "vice")], 29, 31, 52),
        (4, "Tabla 4 (verificador carga)",
         [(179, "producto"), (180, "vice"), (181, "area"), (182, "campo")], 10, 11, 325),
    ]
    rows = []

    def date_cols(hdr_row, start_col):
        """Columnas-fecha desde start_col en hdr_row; salta no-fechas ('(en blanco)', 'Total general')
        y corta tras 3 columnas vacías seguidas (evita cruzar a la tabla siguiente)."""
        out, c, empties = [], start_col, 0
        while empties < 3 and c < start_col + 280:
            v = grid.get((hdr_row, c))
            if v is None:
                empties += 1
            else:
                empties = 0
                d = to_date(v)
                if d is not None:
                    out.append((c, d))
            c += 1
        return out

    for idx, label, dim_cols, hdr_row, r0, r1 in TABLES:
        dcols = date_cols(hdr_row, dim_cols[-1][0] + 1)
        if not dcols:
            continue
        carry = {}
        for r in range(r0, r1 + 1):
            raw = [grid.get((r, dc)) for dc, _ in dim_cols]
            if all(v is None or str(v).strip() == "" for v in raw):
                continue
            # forward-fill de dims con conciencia de subtotales: tras una celda 'Total…',
            # las columnas a su derecha quedan en blanco (no se heredan).
            vals, seen_total = [], False
            for i, v in enumerate(raw):
                sv = s(v)
                if sv:
                    if sv.lower().startswith("total"):
                        seen_total = True
                    carry[i] = sv
                    vals.append(sv)
                else:
                    vals.append(None if seen_total else carry.get(i))
            dims = {name: vals[i] for i, (_, name) in enumerate(dim_cols) if vals[i] is not None}
            for c, d in dcols:
                val = num(grid.get((r, c)))
                if val is None:
                    continue
                rows.append({"tabla_idx": idx, "tabla_label": label,
                             "dims": dims, "fecha": d, "valor": val})
    return {"rows": rows, "tablas": DECLARED}


def _whatsapp_extract(ws):
    """Extractor de 'Reporte Whatsapp' → 12 tablas (valores cacheados TAL COMO APARECEN; NO recalcula;
    decisión usuario 2026-06-30). Contrato genérico [{tabla_idx,tabla_label,dims,fecha,valor}].

    Bloque C–F (cols 3..6) — 6 tablas consolidadas apiladas (idx 1..6; Ecopetrol/Filiales/Upstream ×
      Crudo/Gas/Blancos). 9 filas × 3 métricas D/E/F. dims={segmento, concepto, columna(D/E/F), metrica}.
      ⚠ Los títulos de col C varían entre archivos (mes/trimestre: Junio/Febrero/Octubre, 1Q/4Q); por eso
      las etiquetas son ESTABLES/posicionales y deben coincidir con DECLARED (si no, load_tablas_hoja
      reportaría 0 filas). El segmento se deriva de las filas de subtotal (Ecopetrol/Filiales/Upstream).
    Bloque L–T — 6 tablas por activo (idx 7..12). Las cols Q/R (vacías) separan izquierda L–P (4 métricas
      M/N/O/P) de derecha S–T (2 métricas S/T). 3 secciones (1=Crudo, 2=Gas, 3=Equivalente) ancladas por
      L=='ACTIVOS'; blancos internos NO cortan; cierra en la nota '**…' o el siguiente 'ACTIVOS'.
      dims={activo, columna, metrica} (columna garantiza unicidad cuando metrica se repite: 'Real' diario
      vs acum-mes). La sección Equivalente puede faltar (STD) → T9/T12 quedan vacías (se declaran igual).

    fecha=NULL: la celda 'Producción al:' es poco fiable (en archivos STD viejos trae valores re-calculados
    de otra época); el linaje temporal lo da reporte_id→config_reporte. Celdas #¡REF!/#N/A/(en blanco) →
    s()/num()=None (NOISE) → NO se emiten (se dejan en blanco), según la directriz del dueño. Incluye las
    filas de total/subtotal (decisión del dueño; desvía de la regla de ruido §5#5 a propósito).
    DECLARA siempre las 12 tablas. Anclajes verificados estables en 3 archivos (STD-2023/2024, NEW-2024)."""
    grid, maxr = _p50_grid(ws)
    DECLARED = [
        (1, "T1 PROGRAMA (consolidado)"), (2, "T2 Mes en curso (consolidado)"),
        (3, "T3 Proyección mes (consolidado)"), (4, "T4 Trimestre (consolidado)"),
        (5, "T5 YTD (consolidado)"), (6, "T6 Año (consolidado)"),
        (7, "T7 Crudo por activo (izq L-P)"), (8, "T8 Gas por activo (izq L-P)"),
        (9, "T9 Equivalente por filial (izq L-P)"), (10, "T10 Crudo por activo (der S-T)"),
        (11, "T11 Gas por activo (der S-T)"), (12, "T12 Equivalente por filial (der S-T)"),
    ]
    LBL = {i: l for i, l in DECLARED}
    rows = []
    FECHA = None   # linaje por reporte_id; as-of de la hoja no fiable (ver docstring / §0.c del plan)

    # ===== Bloque C–F (cols 3..6): 6 tablas consolidadas apiladas =====
    HDR_KW = ("real", "plan", "proy", "delta", "programa", "pop")
    SEG_ORDER = ["Ecopetrol", "Filiales", "Upstream"]
    # fila-título = C con texto y D con un rótulo de encabezado (Real/Plan/Proy/Delta/Programa/POP)
    cf_titles = [r for r in range(1, maxr + 1)
                 if s(grid.get((r, 3))) and s(grid.get((r, 4)))
                 and any(k in s(grid.get((r, 4))).lower() for k in HDR_KW)]
    for ti, trow in enumerate(cf_titles[:6]):
        idx = ti + 1                                   # idx posicional ESTABLE (1..6)
        metrics = [(cc, ll, s(grid.get((trow, cc)))) for cc, ll in ((4, "D"), (5, "E"), (6, "F"))]
        metrics = [(cc, ll, m) for cc, ll, m in metrics if m]
        seg_i, r = 0, trow + 1
        while r <= maxr:
            lab = s(grid.get((r, 3)))
            if lab is None:                            # blanco/ruido en col C = fin de la tabla
                break
            lu = lab.upper()
            if lu == "ECOPETROL":
                seg, con, seg_i = "Ecopetrol", "Total", 1
            elif lu == "FILIALES":
                seg, con, seg_i = "Filiales", "Total", 2
            elif lu == "UPSTREAM":
                seg, con = "Upstream", "Total"
            else:
                seg, con = SEG_ORDER[seg_i], lab       # Crudo/Gas/Blancos del segmento actual
            for cc, ll, m in metrics:
                val = num(grid.get((r, cc)))
                if val is None:
                    continue
                rows.append({"tabla_idx": idx, "tabla_label": LBL[idx],
                             "dims": {"segmento": seg, "concepto": con, "columna": ll, "metrica": m},
                             "fecha": FECHA, "valor": val})
            if lu == "UPSTREAM":                       # Upstream cierra la tabla
                break
            r += 1

    # ===== Bloque L–T: secciones ancladas por L=='ACTIVOS' (col 12); Q/R separan izq/der =====
    LEFT = [(13, "M"), (14, "N"), (15, "O"), (16, "P")]      # izquierda L–P (4 métricas)
    RIGHT = [(19, "S"), (20, "T")]                           # derecha S–T (2 métricas)
    SEC = [(7, 10), (8, 11), (9, 12)]                        # orden estable: 1=Crudo, 2=Gas, 3=Equivalente
    sec_hdrs = [r for r in range(1, maxr + 1) if s(grid.get((r, 12))) == "ACTIVOS"]
    for si, hdr in enumerate(sec_hdrs[:3]):
        lidx, ridx = SEC[si]
        lmet = [(cc, ll, s(grid.get((hdr, cc)))) for cc, ll in LEFT]
        rmet = [(cc, ll, s(grid.get((hdr, cc)))) for cc, ll in RIGHT]
        r, blanks = hdr + 1, 0
        while r <= maxr:
            act = s(grid.get((r, 12)))
            if act is None:                            # blanco interno (f14/25/36) NO corta
                blanks += 1
                if blanks >= 3:                        # 3 blancos seguidos = fin de sección
                    break
                r += 1
                continue
            if act.startswith("**") or act == "ACTIVOS":   # nota '**…' o sig. encabezado
                break
            blanks = 0
            for cc, ll, m in lmet:
                val = num(grid.get((r, cc)))
                if val is None:
                    continue
                rows.append({"tabla_idx": lidx, "tabla_label": LBL[lidx],
                             "dims": {"activo": act, "columna": ll, "metrica": m},
                             "fecha": FECHA, "valor": val})
            for cc, ll, m in rmet:
                val = num(grid.get((r, cc)))
                if val is None:
                    continue
                rows.append({"tabla_idx": ridx, "tabla_label": LBL[ridx],
                             "dims": {"activo": act, "columna": ll, "metrica": m},
                             "fecha": FECHA, "valor": val})
            r += 1

    return {"rows": rows, "tablas": DECLARED}

def _mesano_extract(ws):
    """Extractor de 'NEW MES-AÑO' → 13 tablas (cubo fuente de P50/POP; valores cacheados TAL COMO APARECEN;
    NO recalcula; decisión usuario 2026-06-30). Contrato genérico [{tabla_idx,tabla_label,dims,fecha,valor}].

    2 bloques de columnas (anclas verificadas estables en 3 archivos):
      A–O (REAL + PROYECCIÓN), 7 tablas: etiquetas A=producto/concepto, B=vice/empresa; meses=fechas en C–N.
      S–AH (P50 + POP), 6 tablas: etiquetas T=producto, U=vice/empresa (S=índice, se ignora); meses=fechas V–AG.
    Meses por fila-encabezado de fechas (corte en 1ª no-fecha → excluye 'Promedio Año' O/AH, derivado, D-1).
    Incluye filas de total/subtotal (B/U vacío → dims solo con producto/concepto). Tablas de entidad
    (6/10/13): entidad ∈ {GON,GOO,VEX} o 'GRUPO EMPRESARIAL'; el resto de filas del rango se ignora.
    'REAL PROMEDIO MES' (1/2/3) acumula meses según el año del archivo (variación legítima).
    DECLARA siempre las 13 tablas. Excluye GRÁFICAS/META AÑO/serie diaria (otros bloques de columnas)."""
    grid, maxr = _p50_grid(ws)
    DECLARED = [
        (1, "T1 Parámetros calendario (A-O)"), (2, "T2 REAL PROMEDIO MES ECP (A-O)"),
        (3, "T3 REAL PROMEDIO MES Filiales (A-O)"), (4, "T4 PROYECCIÓN AÑO ECP (A-O)"),
        (5, "T5 PROYECCIÓN AÑO Filiales (A-O)"), (6, "T6 POP/PROY EXPLORACIÓN+G.E. (A-O)"),
        (7, "T7 POP Filiales (A-O)"), (8, "T8 P50 ECP (S-AH)"), (9, "T9 P50 Filiales (S-AH)"),
        (10, "T10 P50 EXPLORACIÓN+G.E. (S-AH)"), (11, "T11 POP ECP (S-AH)"),
        (12, "T12 POP Filiales (S-AH)"), (13, "T13 POP EXPLORACIÓN+G.E. (S-AH)"),
    ]
    EXPL = {"GON", "GOO", "VEX"}
    AO_ECP = [(1, "producto"), (2, "vice")];  AO_FIL = [(1, "producto"), (2, "empresa")]
    SAH_ECP = [(20, "producto"), (21, "vice")]; SAH_FIL = [(20, "producto"), (21, "empresa")]
    # (idx, hdr_fechas, fila_ini, fila_fin, col_ini_meses, dim_cols, kind)
    TABLES = [
        (1, 7, 6, 10, 3, [(1, "concepto")], "skiphdr"),
        (2, 13, 14, 35, 3, AO_ECP, ""), (3, 39, 40, 47, 3, AO_FIL, ""),
        (4, 59, 60, 81, 3, AO_ECP, ""), (5, 84, 85, 97, 3, AO_FIL, ""),
        (6, 99, 99, 105, 3, [(1, "a"), (2, "b")], "entidad"),
        (7, 112, 113, 116, 3, [(2, "empresa")], ""),
        (8, 8, 9, 27, 22, SAH_ECP, ""), (9, 30, 31, 43, 22, SAH_FIL, ""),
        (10, 46, 46, 52, 22, [(20, "a"), (21, "b")], "entidad"),
        (11, 59, 60, 81, 22, SAH_ECP, ""), (12, 84, 85, 97, 22, SAH_FIL, ""),
        (13, 99, 99, 105, 22, [(20, "a"), (21, "b")], "entidad"),
    ]
    LBL = {i: l for i, l in DECLARED}
    rows = []
    for idx, hdr, r0, r1, mstart, dim_cols, kind in TABLES:
        months = _p50_contig_months(grid, hdr, mstart)     # corta en 1ª no-fecha (excluye Promedio)
        if not months:
            continue
        for r in range(r0, r1 + 1):
            if kind == "skiphdr" and r == hdr:
                continue
            if kind == "entidad":
                la = s(grid.get((r, dim_cols[0][0]))); lb = s(grid.get((r, dim_cols[1][0])))
                ent = None
                if lb and lb.upper() in EXPL:
                    ent = lb.upper()
                elif la and "GRUPO EMP" in la.upper():
                    ent = "GRUPO EMPRESARIAL"
                elif lb and "GRUPO EMP" in lb.upper():
                    ent = "GRUPO EMPRESARIAL"
                if not ent:
                    continue
                dims = {"entidad": ent}
            else:
                dims = {}
                for col, name in dim_cols:
                    v = s(grid.get((r, col)))
                    if v is not None:
                        dims[name] = v
                if not dims:
                    continue
            for c, d in months:
                val = num(grid.get((r, c)))
                if val is None:
                    continue
                rows.append({"tabla_idx": idx, "tabla_label": LBL[idx],
                             "dims": dims, "fecha": d, "valor": val})
    return {"rows": rows, "tablas": DECLARED}

def _dpp_extract(ws):
    """Extractor de 'Reporte DPP' → 5 tablas MATRICIALES (segmento × métrica), fecha=NULL.
    Reporte comparativo derivado (snapshot de KPIs); NO recalcula (decisión usuario 2026-06-30).
    Contrato extendido: {"rows":[{tabla_idx,tabla_label,dims,fecha,valor}], "tablas":DECLARED}.

    Filas Excel 13-22 = 10 segmentos con jerarquía (ECOPETROL/FILIALES/UPSTREAM). El grupo se
    embebe en dims.fila para evitar colisiones de segmentos repetidos (CRUDO/GAS/BLANCOS x2).
    Columna B = segmento; cabeceras en filas 9-12 (no se emiten).

    REGLA DPP de celdas (acordada 2026-06-30): una celda con valor de error de Excel
    (#¡REF!/#N/A/#DIV0/#VALUE!/#NAME?...) se INSERTA como valor=NULL (celda en blanco, preserva la
    forma de la tabla); una celda VACÍA de verdad se salta. Etiquetas de columna POSICIONALES
    estables (no dependen del mes). Excluye POP 727/725 (cols Y-AE) y filas 74-87 (#REF! total)."""
    DECLARED = [
        (1, "COMPARATIVO DÍA"), (2, "PROYECCIÓN MES"), (3, "CUMPLIMIENTO MES"),
        (4, "PROYECCIÓN AÑO"), (5, "CUMPLIMIENTO AÑO"),
    ]
    # (fila_excel, etiqueta de fila con grupo embebido). Filas fijas verificadas estables en 3 archivos.
    # Las filas total se nombran "TOTAL ..." para que el visor las resalte (chat.js isTotal /^total/i).
    ROWS = [
        (13, "ECOPETROL · CRUDO"), (14, "ECOPETROL · GAS"), (15, "ECOPETROL · BLANCOS"),
        (16, "ECOPETROL · ECP EXPLORACIÓN"), (17, "TOTAL ECOPETROL"),
        (18, "FILIALES · CRUDO"), (19, "FILIALES · GAS"), (20, "FILIALES · BLANCOS"),
        (21, "TOTAL FILIALES"), (22, "TOTAL UPSTREAM"),
    ]
    # (tabla_idx, [(col_index_1based, etiqueta_columna_estable), ...]). Índices: C=3 D=4 E=5 F=6
    # H=8 I=9 K=11 L=12 M=13 O=15 P=16 R=18 S=19 T=20.
    TABLES = [
        (1, [(3, "REAL día anterior"), (4, "PROGRAMA día anterior"),
             (5, "REAL día actual"), (6, "PROGRAMA día actual")]),
        (2, [(8, "REAL mes"), (9, "PROYECCIÓN MES")]),
        (3, [(11, "P50 (META) mes"), (12, "DIFERENCIA mes"), (13, "PRODUCCIÓN NECESARIA mes")]),
        (4, [(15, "REAL año (YTD)"), (16, "PROYECCIÓN AÑO")]),
        (5, [(18, "P50 (META) año"), (19, "DIFERENCIA año"), (20, "PRODUCCIÓN NECESARIA año")]),
    ]
    LBL = {i: l for i, l in DECLARED}
    grid = {}
    for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r > 22:
            break
        for c, v in enumerate(row, start=1):
            grid[(r, c)] = v
    rows = []
    for idx, cols in TABLES:
        for rexcel, fila in ROWS:
            for cidx, clabel in cols:
                raw = grid.get((rexcel, cidx))
                if raw is None:
                    continue
                if str(raw).strip() == "":
                    continue
                # número -> valor; error de Excel (no numérico, no vacío) -> NULL (celda en blanco)
                rows.append({"tabla_idx": idx, "tabla_label": LBL[idx],
                             "dims": {"fila": fila, "columna": clabel},
                             "fecha": None, "valor": num(raw)})
    return {"rows": rows, "tablas": DECLARED}

def _pop_filiales_extract(ws):
    """Extractor de 'POP Filiales y Exploración' → 2 tablas TEMPORALES mensuales (fecha=mes).
    Reporte derivado; NO recalcula (valores cacheados TAL CUAL). Contrato extendido:
    {"rows":[{tabla_idx,tabla_label,dims,fecha,valor}], "tablas":DECLARED}.

    T1 'POP Filiales'    (filas 3-19,  cabecera fila 2):  dims = producto(B) × empresa(C).
    T2 'POP Exploración' (filas 23-26, cabecera fila 22): dims = vr(B) × ger(C).
    Meses contiguos desde col D (4) por la fila-cabecera; _p50_contig_months corta en la 1ª no-fecha
    => EXCLUYE 'Promedio Año' (col P). Incluye subtotales/totales (col C vacía => dims sin 2ª clave;
    cada (producto,empresa)/(vr,ger) es único, 0 colisiones). Fila 1 (días/mes) y col A (índice): metadata."""
    grid, _maxr = _p50_grid(ws)
    DECLARED = [(1, "POP Filiales"), (2, "POP Exploración")]
    LBL = {i: l for i, l in DECLARED}
    # (tabla_idx, hdr_row, fila_ini, fila_fin, nombre_dim1(B), nombre_dim2(C))
    SPECS = [
        (1, 2, 3, 19, "producto", "empresa"),
        (2, 22, 23, 26, "vr", "ger"),
    ]
    rows = []
    for idx, hdr, r0, r1, name1, name2 in SPECS:
        months = _p50_contig_months(grid, hdr, 4)      # D..O ; corta antes de 'Promedio Año'
        if not months:
            continue
        for r in range(r0, r1 + 1):
            a = s(grid.get((r, 2)))                     # B = etiqueta principal
            if a is None:
                continue
            dims = {name1: a}
            b = s(grid.get((r, 3)))                     # C = empresa/ger (ausente en subtotales)
            if b is not None:
                dims[name2] = b
            for c, d in months:
                v = num(grid.get((r, c)))
                if v is None:
                    continue
                rows.append({"tabla_idx": idx, "tabla_label": LBL[idx],
                             "dims": dims, "fecha": d, "valor": v})
    return {"rows": rows, "tablas": DECLARED}

def _calculo_trimestre_extract(ws):
    """Extractor de 'CALCULO DE TRIMESTRE' → 8 tablas (definidas por el usuario). Hoja intermedia de
    cálculo, heterogénea; valores cacheados TAL CUAL (NO recalcula). Contrato extendido.
    T1 (D-I, snapshot matriz): PROGRAMA MES (producto×valor, col 'Programa mes') + FILIALES
       (producto×empresa, col 'Filiales'); fecha=NULL.
    T2/T3/T4 (temporales, fecha=mes): producto×mes / producto×empresa×mes / GRUPO EMPRESA×mes.
       Meses por _p50_contig_months (corta en 'Promedio Año', col O).
    T5/T6/T7/T8 (trimestrales, matriz columna=1Q-4Q, fila=concepto): P50 631 PLAN / P50 631 REAL /
       P50 621,9 PLAN / POP PLAN. En T5/T6 el concepto se repite (bloque ECP filas<62 vs FILIALES
       filas>=62) -> se embebe el bloque en la fila para no perder filas (0 colisiones)."""
    grid, _maxr = _p50_grid(ws)
    DECLARED = [
        (1, "PROGRAMA MES + FILIALES"), (2, "PROYECCIÓN AÑO producto×mes"),
        (3, "PROYECCIÓN AÑO producto×empresa×mes"), (4, "GRUPO EMPRESA×mes"),
        (5, "P50 631 PLAN (trimestres)"), (6, "P50 631 REAL (trimestres)"),
        (7, "P50 621,9 PLAN (trimestres)"), (8, "POP PLAN (trimestres)"),
    ]
    LBL = {i: l for i, l in DECLARED}
    rows = []

    def emit(idx, dims, fecha, raw):
        v = num(raw)
        if v is None:
            return
        rows.append({"tabla_idx": idx, "tabla_label": LBL[idx],
                     "dims": dims, "fecha": fecha, "valor": v})

    # --- T1 (cols D-I, filas 8-22) — matriz snapshot, fecha=NULL ---
    for r in range(8, 23):
        a = s(grid.get((r, 4)))                       # D: producto (PROGRAMA MES)
        if a is not None:
            emit(1, {"fila": a, "columna": "Programa mes"}, None, grid.get((r, 5)))   # E
        gl = s(grid.get((r, 7)))                      # G: producto (FILIALES)
        if gl is not None:
            emp = s(grid.get((r, 8)))                 # H: empresa
            fila = f"{gl} · {emp}" if emp else gl
            emit(1, {"fila": fila, "columna": "Filiales"}, None, grid.get((r, 9)))     # I

    # --- T2/T3/T4 — temporales (fecha=mes; meses desde col C=3) ---
    for r in range(25, 32):                           # T2: producto × mes
        a = s(grid.get((r, 1)))
        if a is None:
            continue
        for c, d in _p50_contig_months(grid, 24, 3):
            emit(2, {"producto": a}, d, grid.get((r, c)))
    for r in range(36, 49):                           # T3: producto × empresa × mes
        a = s(grid.get((r, 1)))
        if a is None:
            continue
        dims = {"producto": a}
        b = s(grid.get((r, 2)))
        if b is not None:
            dims["empresa"] = b
        for c, d in _p50_contig_months(grid, 35, 3):
            emit(3, dims, d, grid.get((r, c)))
    for r in range(51, 52):                           # T4: GRUPO EMPRESA × mes
        a = s(grid.get((r, 1)))
        if a is None:
            continue
        for c, d in _p50_contig_months(grid, 50, 3):
            emit(4, {"concepto": a}, d, grid.get((r, c)))

    # --- T5/T6/T7/T8 — trimestrales (matriz columna=1Q-4Q, fila=concepto[+bloque en T5/T6]) ---
    QLABELS = ["1Q", "2Q", "3Q", "4Q"]
    # (tabla_idx, col_etiqueta, [cols_trimestre], rango_filas, usa_bloque)
    QSPECS = [
        (5, 1, [2, 3, 4, 5], range(55, 67), True),
        (6, 8, [9, 10, 11, 12], range(55, 69), True),
        (7, 1, [2, 3, 4, 5], range(77, 83), False),
        (8, 1, [2, 3, 4, 5], range(88, 95), False),
    ]
    for idx, lcol, qcols, rng, usa_bloque in QSPECS:
        # [v2] sufijar el bloque SOLO a las etiquetas que se repiten dentro de la tabla
        # (CRUDO/GAS/BLANCOS) -> VDP/FILIALES/TOTAL/UPSTREAM/VEX... quedan limpias. Sin imports nuevos.
        labels = [s(grid.get((r, lcol))) for r in rng]
        dups = {x for x in labels if x is not None and labels.count(x) > 1} if usa_bloque else set()
        for r in rng:
            a = s(grid.get((r, lcol)))
            if a is None:
                continue
            fila = f"{a} ({'ECP' if r < 62 else 'FILIALES'})" if a in dups else a
            for c, q in zip(qcols, QLABELS):
                emit(idx, {"fila": fila, "columna": q}, None, grid.get((r, c)))

    return {"rows": rows, "tablas": DECLARED}

def _inicio_extract(ws):
    """Extractor de la hoja 'INICIO' → 1 tabla: 'REAL PROMEDIO MES (YTD) Filiales' (único dato a mano).
    Temporal mensual (fecha=mes, YTD), dims producto×empresa. Se ANCLA POR TÍTULO (no por filas fijas)
    porque la tabla se desplaza entre archivos NEW (~fila 34) y STD (~fila 38). Meses dinámicos por el
    header (corta en 'Promedio Año'/no-fecha); valores YTD hasta el corte. El resto de INICIO
    (parámetros/lookups de setup) NO se ingiere. Contrato extendido {"rows":..., "tablas":DECLARED}."""
    grid, maxr = _p50_grid(ws)
    DECLARED = [(1, "REAL PROMEDIO MES (YTD) Filiales")]
    trow = None
    for r in range(1, maxr + 1):
        v = s(grid.get((r, 1)))
        if v and v.upper().startswith("REAL PROMEDIO MES (YTD) FILIALES"):
            trow = r
            break
    if trow is None:
        return {"rows": [], "tablas": DECLARED}
    hdr = trow + 1
    months = _p50_contig_months(grid, hdr, 3)          # meses desde col C (3); corta en no-fecha
    rows = []
    r = hdr + 1
    while r <= maxr and s(grid.get((r, 1))) is not None:
        a = s(grid.get((r, 1)))
        dims = {"producto": a}
        b = s(grid.get((r, 2)))
        if b is not None:
            dims["empresa"] = b
        for c, d in months:
            v = num(grid.get((r, c)))
            if v is not None:
                rows.append({"tabla_idx": 1, "tabla_label": "REAL PROMEDIO MES (YTD) Filiales",
                             "dims": dims, "fecha": d, "valor": v})
        r += 1
    return {"rows": rows, "tablas": DECLARED}

def _td_datos_dia_extract(ws):
    """Extractor de la Tabla Dinámica 'TD_datos_dia' → 1 tabla larga temporal (fecha=día), grano DETALLE
    (filas-hoja × ECOPETROL/SOCIOS × día). EXCLUYE subtotales (filas 'Total …' y columnas de total por
    fecha/gran total: derivados). Cabeceras ANCLADAS POR CONTENIDO (el layout difiere NEW/STD):
      fila 19 = FECHA (ffill desde la última fecha real); fila 20 = medida 'Suma de <X>' (ffill;
      X ∈ VOLDISMEZ/VOL_ESTIMADO/PROMEDIO); fila 21 = GRUPOPROD (ECOPETROL/SOCIOS) → columna de detalle.
    Campos de fila A-E = TIPOPRODUCTO/VICE/ACTIVOS/GRUPO1/FUENTE (ffill de niveles padre). '(en blanco)'
    se PRESERVA como categoría real (NO usar s() en los campos de fila). dims = 5 niveles + grupoprod +
    medida. NO usar _p50_grid (corta en fila 250; TD tiene ~590). Verificado 0 colisiones en 3 archivos."""
    DECLARED = [(1, "TD_datos_dia (detalle diario)")]
    # grid COMPLETO (sin cap de filas)
    grid = {}
    maxr = maxc = 0
    for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for c, v in enumerate(row, start=1):
            if v is not None and str(v).strip() != "":
                grid[(r, c)] = v
                if r > maxr:
                    maxr = r
                if c > maxc:
                    maxc = c

    def rl(v):                                  # etiqueta de fila: preserva '(en blanco)'; '' / None = ausente
        if v is None:
            return None
        t = str(v).strip()
        return None if t == "" else t

    # --- columnas de detalle: grupoprod ∈ {ECOPETROL,SOCIOS} con fecha y medida vigentes ---
    detail = []                                 # [(col, fecha, medida, grupoprod)]
    cur_date = cur_meas = None
    for c in range(6, maxc + 1):
        d = to_date(grid.get((19, c)))
        if d:
            cur_date = d
        m = s(grid.get((20, c)))
        if m and m.lower().startswith("suma de "):
            cur_meas = m[8:].strip()
        gp = s(grid.get((21, c)))
        if gp in ("ECOPETROL", "SOCIOS") and cur_date and cur_meas:
            detail.append((c, cur_date, cur_meas, gp))
    if not detail:
        return {"rows": [], "tablas": DECLARED}

    # --- filas de datos (desde 22): ffill niveles padre; saltar 'Total …' (subtotales) ---
    ROWF = ["tipoproducto", "vice", "activos", "grupo1", "fuente"]
    rows = []
    ff = [None] * 5
    for r in range(22, maxr + 1):
        vals = [rl(grid.get((r, c))) for c in range(1, 6)]
        if any(v and v.lower().startswith("total") for v in vals):
            continue                            # subtotal/gran total → derivado, se excluye
        for i in range(5):
            if vals[i] is not None:
                ff[i] = vals[i]
        base = {ROWF[i]: ff[i] for i in range(5) if ff[i] is not None}
        for c, d, meas, gp in detail:
            v = num(grid.get((r, c)))
            if v is None:
                continue
            dims = dict(base)
            dims["grupoprod"] = gp
            dims["medida"] = meas
            rows.append({"tabla_idx": 1, "tabla_label": "TD_datos_dia (detalle diario)",
                         "dims": dims, "fecha": d, "valor": v})
    return {"rows": rows, "tablas": DECLARED}

def _datos_mes_extract(ws):
    """Extractor de la Tabla Dinámica MENSUAL 'DATOS_MES' (pivot de BDP_datos_mes) → 1 tabla larga
    temporal (fecha=fin de mes), grano DETALLE (escenario × producto × vice × activos × area × campo).
    EXCLUYE subtotales (filas 'Total …' y la columna 'Total general'). Cabeceras ANCLADAS POR CONTENIDO
    (el layout difiere NEW/STD): la fila de cabecera es la que tiene 'ESCENARIO' en col A; cols A-F = los
    6 niveles de fila; desde col G las fechas mensuales (enteros yyyymmdd). Medida única 'BPDEQ_M' y
    GRUPOPROD fijo=ECOPETROL son filtros del pivot (no se modelan como dims). '(en blanco)' se PRESERVA
    como categoría real (NO usar s() en los campos de fila). NO usar _p50_grid (corta en fila 250;
    DATOS_MES llega a ~1275). Verificado 0 colisiones en 3 archivos."""
    DECLARED = [(1, "DATOS_MES (detalle mensual)")]
    # grid COMPLETO (sin cap de filas)
    grid = {}
    maxr = maxc = 0
    for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for c, v in enumerate(row, start=1):
            if v is not None and str(v).strip() != "":
                grid[(r, c)] = v
                if r > maxr:
                    maxr = r
                if c > maxc:
                    maxc = c

    def rl(v):                                  # etiqueta de fila: preserva '(en blanco)'; '' / None = ausente
        if v is None:
            return None
        t = str(v).strip()
        return None if t == "" else t

    # --- fila de cabecera: la que tiene 'ESCENARIO' en col A ---
    hdr = None
    for r in range(1, maxr + 1):
        if (s(grid.get((r, 1))) or "").upper() == "ESCENARIO":
            hdr = r
            break
    if hdr is None:
        return {"rows": [], "tablas": DECLARED}

    # --- columnas de fecha (desde col 7); no-fecha (p.ej. 'Total general') → ignorada ---
    datecols = []                               # [(col, fecha)]
    for c in range(7, maxc + 1):
        d = to_date(grid.get((hdr, c)))
        if d:
            datecols.append((c, d))
    if not datecols:
        return {"rows": [], "tablas": DECLARED}

    # --- filas de datos: ffill niveles padre; saltar 'Total …' (subtotales) ---
    ROWF = ["escenario", "producto", "vice", "activos", "area", "campo"]
    rows = []
    ff = [None] * 6
    for r in range(hdr + 1, maxr + 1):
        vals = [rl(grid.get((r, c))) for c in range(1, 7)]
        if any(v and v.lower().startswith("total") for v in vals):
            continue                            # subtotal/gran total → derivado, se excluye
        for i in range(6):
            if vals[i] is not None:
                ff[i] = vals[i]
        base = {ROWF[i]: ff[i] for i in range(6) if ff[i] is not None}
        for c, d in datecols:
            v = num(grid.get((r, c)))
            if v is None:
                continue
            rows.append({"tabla_idx": 1, "tabla_label": "DATOS_MES (detalle mensual)",
                         "dims": dict(base), "fecha": d, "valor": v})
    return {"rows": rows, "tablas": DECLARED}

def _bdp_datos_dia_extract(ws):
    """Extractor de la hoja RAW plana 'BDP_datos_dia' (solo archivos NEW) → 1 tabla larga temporal
    (fecha=día) por UNPIVOT de las 5 medidas (VOLUMEN/PORCENTAJE/VOLDISMEZ/VOL_ESTIMADO/PROMEDIO).
    dims = columnas descriptivas (excluye FECHA/MES/AÑO y las medidas) + 'medida'. Cabecera en la fila 1
    (mapa nombre→índice, robusto a reordenamiento de columnas). fecha=FECHA (entero yyyymmdd). Emite una
    fila por cada (registro × medida con valor no nulo), conservando 0 reales (~40K filas). Grano RAW
    completo (D3). Cada fila raw es atómica → se usa s() (no hay subtotales ni ffill). Redundante con
    fact_produccion_dia_ecp y con el pivot TD_datos_dia; aquí al grano fino para el visor."""
    DECLARED = [(1, "BDP_datos_dia (detalle diario RAW)")]
    it = ws.iter_rows(values_only=True)
    header = None
    for row in it:
        header = row
        break
    if header is None:
        return {"rows": [], "tablas": DECLARED}
    name_at = {}                                 # col-idx -> nombre de cabecera (no vacío)
    for i, h in enumerate(header):
        hn = s(h)
        if hn is not None:
            name_at[i] = hn
    upper = {hn.upper(): i for i, hn in name_at.items()}
    MEAS = ["VOLUMEN", "PORCENTAJE", "VOLDISMEZ", "VOL_ESTIMADO", "PROMEDIO"]
    if "FECHA" not in upper or not any(m in upper for m in MEAS):
        return {"rows": [], "tablas": DECLARED}
    fecha_i = upper["FECHA"]
    EXCLUDE = {"FECHA", "MES", "AÑO", "ANO"} | set(MEAS)
    dim_idx = [(i, name_at[i].lower()) for i in sorted(name_at)
               if name_at[i].upper() not in EXCLUDE]
    meas_idx = [(upper[m], m) for m in MEAS if m in upper]
    rows = []
    for row in it:
        d = to_date(row[fecha_i]) if fecha_i < len(row) else None
        if d is None:
            continue
        base = {}
        for i, key in dim_idx:
            if i < len(row):
                val = s(row[i])
                if val is not None:
                    base[key] = val
        for mi, mname in meas_idx:
            v = num(row[mi]) if mi < len(row) else None
            if v is None:
                continue
            dims = dict(base)
            dims["medida"] = mname
            rows.append({"tabla_idx": 1, "tabla_label": "BDP_datos_dia (detalle diario RAW)",
                         "dims": dims, "fecha": d, "valor": v})
    return {"rows": rows, "tablas": DECLARED}

def _bdp_datos_mes_extract(ws):
    """Extractor de la hoja RAW mensual plana 'BDP_datos_mes' (solo archivos NEW) → 1 tabla larga temporal
    (fecha=fin de mes), 1 FILA POR REGISTRO (NO se despliegan las 10 medidas → 314.952 filas, no el unpivot).
    valor = BPDEQ_M (barriles/día equivalente mensual, la misma medida del pivot DATOS_MES). dims = 46
    columnas descriptivas (excluye FECHA/MES/AÑO y las 10 medidas). Cabecera en la fila 1 (mapa nombre→índice,
    robusto a reordenamiento). fecha=FECHA (entero yyyymmdd). rl() preserva '(en blanco)' y pasa fechas
    descriptivas a ISO → 0 colisiones (dims+fecha únicos). Cubo multi-año (97 meses 2020-2028, 4 escenarios).
    Grano RAW completo (D3); redundante con fact_produccion_mes_ecp y el pivot DATOS_MES. NO usar _p50_grid."""
    DECLARED = [(1, "BDP_datos_mes (detalle mensual RAW)")]
    it = ws.iter_rows(values_only=True)
    header = None
    for row in it:
        header = row
        break
    if header is None:
        return {"rows": [], "tablas": DECLARED}
    name_at = {}                                 # col-idx -> nombre de cabecera (no vacío)
    for i, h in enumerate(header):
        hn = s(h)
        if hn is not None:
            name_at[i] = hn
    upper = {hn.upper(): i for i, hn in name_at.items()}
    MEAS = ["VOLUMEN", "PORCENTAJE", "VOLDISMEZ", "BPD_M", "BPDA_AC", "BPDAC_5",
            "BPD_A", "BPDEQ_M", "BLSEQ", "BPDEQ_A"]
    if "FECHA" not in upper or "BPDEQ_M" not in upper:
        return {"rows": [], "tablas": DECLARED}
    fecha_i = upper["FECHA"]
    val_i = upper["BPDEQ_M"]
    EXCLUDE = {"FECHA", "MES", "AÑO", "ANO"} | set(MEAS)
    dim_idx = [(i, name_at[i].lower()) for i in sorted(name_at)
               if name_at[i].upper() not in EXCLUDE]

    def rl(v):                                  # dim: preserva '(en blanco)'; fecha→ISO; '' / None = ausente
        if v is None:
            return None
        if isinstance(v, (dt.date, dt.datetime)):
            return v.isoformat()
        t = str(v).strip()
        return None if t == "" else t

    rows = []
    for row in it:
        d = to_date(row[fecha_i]) if fecha_i < len(row) else None
        if d is None:
            continue
        v = num(row[val_i]) if val_i < len(row) else None
        dims = {}
        for i, key in dim_idx:
            if i < len(row):
                val = rl(row[i])
                if val is not None:
                    dims[key] = val
        rows.append({"tabla_idx": 1, "tabla_label": "BDP_datos_mes (detalle mensual RAW)",
                     "dims": dims, "fecha": d, "valor": v})
    return {"rows": rows, "tablas": DECLARED}

def _bdp_programa_extract(ws):
    """Extractor de la hoja RAW plana 'BDP_Programa' (solo archivos NEW) → 1 tabla larga temporal
    (fecha = fecha del programa), 1 FILA POR REGISTRO (13.822; SIN unpivot de las 3 medidas). valor = Volumen
    (cuota ECP programada). dims = las 12 columnas restantes, incluyendo Produccion_total y Part_ECP
    preservadas como dims (para no perder ninguna columna: decisión del usuario '13.823x14 = la ingesta').
    Cabecera en la fila 1 (mapa nombre→índice, robusto a reordenamiento); clave de dim = header en minúsculas
    con espacios→'_' (p.ej. 'Fecha Version' -> 'fecha_version'). fecha=Fecha (string yyyymmdd). rl() pasa fechas
    descriptivas a ISO y descarta ruido/'(en blanco)'. 0 colisiones (dims+fecha únicos). Tabla plana atómica
    (0 subtotales verificados). Grano RAW completo (D3); 3.er destino, redundante con bronze.bdp_programa y la
    estrella core.fact_programa_ecp."""
    DECLARED = [(1, "BDP_Programa (programa RAW)")]
    it = ws.iter_rows(values_only=True)
    header = None
    for row in it:
        header = row
        break
    if header is None:
        return {"rows": [], "tablas": DECLARED}
    name_at = {}                                 # col-idx -> nombre de cabecera (no vacío)
    for i, h in enumerate(header):
        hn = s(h)
        if hn is not None:
            name_at[i] = hn
    upper = {hn.upper(): i for i, hn in name_at.items()}
    if "FECHA" not in upper or "VOLUMEN" not in upper:
        return {"rows": [], "tablas": DECLARED}
    fecha_i = upper["FECHA"]
    val_i = upper["VOLUMEN"]
    # dims = todas las columnas con cabecera excepto FECHA (→fecha) y VOLUMEN (→valor).
    # Produccion_total y Part_ECP quedan como dims (no se pierde ninguna de las 14 columnas).
    # clave = header.lower() con espacios→'_' (limpia 'Fecha Version' -> 'fecha_version').
    EXCLUDE = {"FECHA", "VOLUMEN"}
    dim_idx = [(i, name_at[i].lower().replace(" ", "_")) for i in sorted(name_at)
               if name_at[i].upper() not in EXCLUDE]

    def rl(v):                                  # dim: fecha→ISO; ruido/'(en blanco)'/'' / None = ausente
        if v is None:
            return None
        if isinstance(v, (dt.date, dt.datetime)):
            return v.isoformat()
        t = str(v).strip()
        return None if t in NOISE else t

    rows = []
    for row in it:
        d = to_date(row[fecha_i]) if fecha_i < len(row) else None
        if d is None:
            continue
        v = num(row[val_i]) if val_i < len(row) else None
        dims = {}
        for i, key in dim_idx:
            if i < len(row):
                val = rl(row[i])
                if val is not None:
                    dims[key] = val
        rows.append({"tabla_idx": 1, "tabla_label": "BDP_Programa (programa RAW)",
                     "dims": dims, "fecha": d, "valor": v})
    return {"rows": rows, "tablas": DECLARED}


def _reporte_president_extract(ws):
    """Extractor de 'REPORTE_PRESIDENT' → tabla comparativa de producción (Δ vs plan), grano ENTIDAD×MEDIDA.
    Contrato genérico [{tabla_idx,tabla_label,dims,fecha,valor}]; dims={entidad, medida}; fecha=NULL
    (snapshot del corte; linaje por reporte_id, precedente 'Reporte Whatsapp').

    Verificado en 3 archivos (ene/mar/may 2026): fila de encabezado = 37 (anclada por 'Base P50').
    Dos bloques que comparten la fila de encabezado, cada uno con su PROPIA columna de etiqueta de entidad:
      BLOQUE DÍA  (B:E)  B=entidad · C='Real día' · D='Programa día' · E='Delta'
      BLOQUE MES  (G:M)  G=entidad · H='Real Mes <mes>' · I='Proy. Mes <mes>' · J='Base P50'
                          · K='Delta' · L='P50'/'Reto NNNK' (compromiso) · M='Delta'
    Entidades (filas 38-43): Crudo, Gas, Blancos, Ecopetrol (=Σ productos), Filiales, Upstream.
    Escala = kbpe (mundo corporativo P50; NO la del fact diario). 🔑 Trae BLANCOS con Real/Proy/P50
    (que la fuente 'REAL PROMEDIO MES' de NEW MES-AÑO NO tenía) y el 'compromiso' (col L = Reto cuando
    difiere del P50; = P50 cuando no hay stretch ese periodo).

    Medidas por POSICIÓN (el sufijo de mes 'Mar/Dic/May' y 'P50'/'Reto NNNK' del encabezado VARÍAN entre
    archivos → NO se usa el texto crudo como clave). El bloque DÍA suele venir en #REF!/#N/A en varios
    archivos → num() lo descarta y esa medida no emite fila (correcto). El bloque MES es el fiable.

    Tablas: entrada 1 tabla numérica (2 bloques) → salida 2 (día, mes). EXCLUIDO (narrativa, no numérico,
    avisado al usuario 2026-07-27): 'Principales eventos' (r45+) y 'Actividades programadas' (r54+) — texto
    libre. El gráfico 'Seguimiento de producción' (r7) referencia datos de OTRA hoja (r8-32 vacías)."""
    grid = {}
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=50, max_col=14, values_only=True), 1):
        for c, v in enumerate(row, 1):
            if v is not None and str(v).strip() != "":
                grid[(r, c)] = v

    DECLARED = [(1, "Tabla 1 (Producción día)"), (2, "Tabla 2 (Producción mes)")]
    rows = []

    def find_cell(pred):
        for (r, c), v in grid.items():
            if isinstance(v, str) and pred(v.strip().lower()):
                return r, c
        return None

    anchor = find_cell(lambda t: t == "base p50")          # ancla estable → J37
    if not anchor:
        return {"rows": rows, "tablas": DECLARED}
    hdr, jcol = anchor
    dia = find_cell(lambda t: t in ("real día", "real dia"))
    dcol = dia[1] if (dia and dia[0] == hdr) else None      # C37 (mismo hdr) o None

    gcol = jcol - 3                                         # etiqueta de entidad del bloque MES (G = J-3)
    ent_rows = []
    r = hdr + 1
    while s(grid.get((r, gcol))):                           # hasta la 1ª fila sin entidad (col G, siempre limpia)
        ent_rows.append((r, s(grid.get((r, gcol)))))
        r += 1

    # (offset_relativo, medida) — MES relativo a jcol (Base P50); DÍA relativo a dcol (Real día)
    MES = [(-2, "real_mes"), (-1, "proy_mes"), (0, "base_p50"),
           (1, "delta_p50"), (2, "compromiso"), (3, "delta_compromiso")]
    DIA = [(0, "real_dia"), (1, "programa_dia"), (2, "delta_dia")]

    for r, ent in ent_rows:
        for off, med in MES:
            v = num(grid.get((r, jcol + off)))
            if v is None:
                continue
            rows.append({"tabla_idx": 2, "tabla_label": "Tabla 2 (Producción mes)",
                         "dims": {"entidad": ent, "medida": med}, "fecha": None, "valor": v})
        if dcol:
            ent_d = s(grid.get((r, dcol - 1))) or ent       # etiqueta del bloque día (col B)
            for off, med in DIA:
                v = num(grid.get((r, dcol + off)))
                if v is None:
                    continue
                rows.append({"tabla_idx": 1, "tabla_label": "Tabla 1 (Producción día)",
                             "dims": {"entidad": ent_d, "medida": med}, "fecha": None, "valor": v})
    return {"rows": rows, "tablas": DECLARED}

# === Registro de hojas modeladas: (regex_nombre_hoja, extractor) ===
# Agregar una hoja nueva en el futuro = una línea aquí + su función _xxx_extract(ws).
HOJAS_MODELADAS = [
    # Prefijo (no exige el nombre completo): Excel trunca los nombres de hoja a 31 chars y
    # "P50 Quemado 2024 ECP y Filiales" mide exactamente 31 → cualquier copia puede quedar como
    # "...ECP y Fili". Por eso matcheamos por prefijo, tolerante a la truncación.
    (re.compile(r"(?i)^P50 Quemado \d{4} ECP y Fili"), _p50_extract),
    # Producción filiales: 3 tablas diarias REAL/PROGRAMA/PROYECCIÓN (nombre 19 chars, sin truncar)
    (re.compile(r"(?i)^Producci[oó]n filiales"), _filiales_extract),
    # (Bitacora): 3 tablas diarias TIPOPRODUCTO×VICE (REAL/PROGRAMA/PROYECCIÓN). El nombre real es "(Bitacora)".
    (re.compile(r"(?i)^\(?\s*bit[aá]cora"), _bitacora_extract),
    # P50 Acumulado: 2 tablas de promedios P50 acumulados (producto×mes). Valores ya calculados,
    # se ingieren tal cual (sin recalcular). Nombre 14 chars → sin truncar.
    (re.compile(r"(?i)^P50 Acumulado"), _p50_acum_extract),
    # REPORTE_PRESIDENT: tabla comparativa Δ vs plan (entidad×medida), bloques día+mes. 🔑 Única fuente
    # con BLANCOS Real/Proy/P50 + el 'compromiso' (Reto) del presidente, en escala kbpe corporativa.
    (re.compile(r"(?i)^REPORTE_PRESIDENT$"), _reporte_president_extract),
    # PROGRAMA: 4 tablas (3 pivots + 1 no-pivot) producto/area/campo × fecha. Valores ya calculados,
    # se ingieren tal cual (sin recalcular desde BDP_Programa).
    (re.compile(r"(?i)^PROGRAMA$"), _programa_extract),
    # Reporte Whatsapp: 12 tablas (6 consolidadas C-F + 6 por activo L-T, izq/der separadas por Q-R).
    # Valores ya calculados, se ingieren tal cual. fecha=NULL (as-of de la hoja no fiable en STD viejos).
    # Etiquetas estables (los títulos de mes/trimestre cambian entre archivos). Nombre 16 chars → sin truncar.
    (re.compile(r"(?i)^Reporte\s+Whatsapp"), _whatsapp_extract),
    # NEW MES-AÑO: 13 tablas (cubo fuente P50/POP). A-O REAL/PROYECCIÓN (7) + S-AH P50/POP (6).
    # Valores ya calculados, se ingieren tal cual. Meses por encabezado de fecha; excluye 'Promedio Año'.
    (re.compile(r"(?i)^NEW MES-?A[ÑN]O"), _mesano_extract),
    # Reporte DPP: 5 tablas matriciales (segmento × métrica), comparativo derivado. fecha=NULL.
    # Errores de Excel se ingieren como valor=NULL (celda en blanco, no se descarta). Nombre 11 chars.
    (re.compile(r"(?i)^Reporte\s+DPP"), _dpp_extract),
    # POP Filiales y Exploración: 2 tablas temporales mensuales (T1 Producto×Empresa, T2 VR×GER).
    # Meses por cabecera (corta en 'Promedio Año'). Valores ya calculados, se ingieren tal cual.
    (re.compile(r"(?i)^POP Filiales y Explora"), _pop_filiales_extract),
    # CALCULO DE TRIMESTRE: hoja intermedia, 8 tablas heterogéneas (1 snapshot + 3 temporales + 4
    # trimestrales 1Q-4Q). Valores ya calculados, se ingieren tal cual. Datos parcialmente duplicados
    # de otras hojas (autorizado por el usuario / D3).
    (re.compile(r"(?i)^C[AÁ]LCULO DE TRIMESTRE"), _calculo_trimestre_extract),
    # INICIO: hoja de setup; se ingiere SOLO la tabla 'REAL PROMEDIO MES (YTD) Filiales' (único dato a
    # mano), anclada por título (se desplaza entre NEW/STD). producto×empresa×mes, YTD. Dato redundante.
    (re.compile(r"(?i)^INICIO$"), _inicio_extract),
    # TD_datos_dia: Tabla Dinámica diaria (pivot). Grano detalle (5 niveles × ECOPETROL/SOCIOS × medida
    # × día); subtotales excluidos. Cabeceras ancladas por contenido (layout NEW/STD difiere). Volumen alto.
    (re.compile(r"(?i)^TD_datos_dia$"), _td_datos_dia_extract),
    # DATOS_MES: Tabla Dinámica MENSUAL (pivot de BDP_datos_mes). Grano detalle (6 niveles × mes);
    # subtotales excluidos. Cabeceras por contenido (ancla 'ESCENARIO'); '(en blanco)' preservado. Volumen alto.
    (re.compile(r"(?i)^DATOS_MES$"), _datos_mes_extract),
    # BDP_datos_dia: hoja RAW plana (solo archivos NEW). Unpivot de 5 medidas -> ~40K filas largas
    # (fecha=día). Grano RAW completo; redundante con fact_produccion_dia_ecp y el pivot TD_datos_dia.
    (re.compile(r"(?i)^BDP_datos_dia$"), _bdp_datos_dia_extract),
    # BDP_datos_mes: hoja RAW mensual plana (solo archivos NEW). 1 fila por registro (~314.952 filas),
    # valor=BPDEQ_M; dims=46 descriptivas. Cubo multi-año (2020-2028, 4 escenarios). Redundante con
    # fact_produccion_mes_ecp y el pivot DATOS_MES. Volumen alto (~8x PROGRAMA).
    (re.compile(r"(?i)^BDP_datos_mes$"), _bdp_datos_mes_extract),
    # BDP_Programa: hoja RAW plana del PROGRAMA (solo archivos NEW). 1 fila por registro (13.822 filas),
    # valor=Volumen; dims=12 (incl. Produccion_total y Part_ECP preservadas). Horizonte a futuro (fecha del
    # programa). 3.er destino, redundante con core.fact_programa_ecp y bronze.bdp_programa. Regex disjunto de ^PROGRAMA$.
    (re.compile(r"(?i)^BDP_Programa$"), _bdp_programa_extract),
]

def load_tablas_hoja(conn, wb, reporte_id):
    """Por cada hoja del registro presente en el libro: extrae filas (contrato genérico), deduplica por
    (tabla_idx, dims, fecha) last-wins, reemplaza en core.fact_tabla_hoja (DELETE+INSERT por hoja) y
    devuelve resúmenes [{hoja, tablas:[{tabla_idx,tabla_label,filas}], total}] para el front."""
    resumen = []
    for pat, extractor in HOJAS_MODELADAS:
        hoja = next((sh for sh in wb.sheetnames if pat.match(sh)), None)
        if not hoja:
            continue
        res = extractor(wb[hoja])
        if isinstance(res, dict):                       # contrato extendido: filas + tablas declaradas
            filas = res.get("rows", [])
            declared = res.get("tablas")                # [(idx,label),...] o None
        else:
            filas, declared = res, None
        by_key = {}
        for f in filas:
            k = (f["tabla_idx"], json.dumps(f["dims"], sort_keys=True, ensure_ascii=False), f["fecha"])
            by_key[k] = f
        out = list(by_key.values())
        conn.execute(sa.text("DELETE FROM core.fact_tabla_hoja WHERE reporte_id=:r AND hoja=:h"),
                     {"r": reporte_id, "h": hoja})
        if out:
            _ins_th = sa.text("""
                INSERT INTO core.fact_tabla_hoja (reporte_id, hoja, tabla_idx, tabla_label, dims, fecha, valor)
                VALUES (:r, :h, :idx, :label, CAST(:dims AS jsonb), :fecha, :valor)
            """)
            for _i in range(0, len(out), CHUNK):
                conn.execute(_ins_th,
                    [{"r": reporte_id, "h": hoja, "idx": f["tabla_idx"], "label": f["tabla_label"],
                      "dims": json.dumps(f["dims"], ensure_ascii=False),
                      "fecha": f["fecha"], "valor": f["valor"]} for f in out[_i:_i + CHUNK]])
        cont = {}
        for f in out:
            kk = (f["tabla_idx"], f["tabla_label"])
            cont[kk] = cont.get(kk, 0) + 1
        if declared:                                    # incluir todas las declaradas (aunque 0 filas)
            tablas = [{"tabla_idx": i, "tabla_label": l, "filas": cont.get((i, l), 0)} for i, l in declared]
        else:
            tablas = [{"tabla_idx": i, "tabla_label": l, "filas": n} for (i, l), n in sorted(cont.items())]
        resumen.append({"hoja": hoja, "tablas": tablas, "total": len(out)})
    return resumen

def ingerir_archivo(path: Path, progress_cb=None) -> ResultadoIngesta:
    """Ingesta completa de un .xlsm (NEW o STD). Devuelve ResultadoIngesta con filas_por_tabla.
    Si progress_cb se provee, se invoca por hoja con dicts {"tipo":"hoja","hoja":..,"estado":..,..}."""
    def _emit(ev):
        if progress_cb:
            try: progress_cb(ev)
            except Exception: pass
    wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    raw = tiene_raw(set(wb.sheetnames))
    filas: dict[str, int] = {}
    _emit({"tipo": "inicio", "archivo": path.name, "hojas": list(wb.sheetnames),
           "total": len(wb.sheetnames), "tipo_archivo": "NEW" if raw else "STD"})

    with get_engine().begin() as conn:
        reporte_id, fecha_rep = get_reporte(conn, path, raw)

        # caches de dims
        vice  = DimCache(conn, "core.dim_vicepresidencia", "vice_id", "codigo")
        socio = DimCache(conn, "core.dim_socio", "socio_id", "nombre")
        conc  = DimCache(conn, "core.dim_concepto", "concepto_id", "nombre")
        tipo  = DimCache(conn, "core.dim_tipo_producto", "tipo_producto_id", "nombre")
        esc   = DimCache(conn, "core.dim_escenario", "escenario_id", "nombre")
        proc  = DimCache(conn, "core.dim_proceso", "proceso_id", "nombre")
        emp   = DimCache(conn, "core.dim_empresa", "empresa_id", "nombre")
        treg  = DimCache(conn, "core.dim_tipo_registro", "tipo_id", "nombre")

        # ---- BRONZE tipado (raw) ----
        if raw:
            for sheet, table, cols in [("BDP_datos_dia", "bdp_datos_dia", BZ_DIA),
                                       ("BDP_datos_mes", "bdp_datos_mes", BZ_MES),
                                       ("BDP_Programa",  "bdp_programa",  BZ_PRG)]:
                _emit({"tipo": "hoja", "hoja": sheet, "estado": "procesando"})
                n = land_bronze_typed(conn, wb[sheet], table, cols, reporte_id)
                _log_ingesta(conn, reporte_id, sheet, f"bronze.{table}", n, n)
                filas[f"bronze.{table}"] = n
                log.info("ingesta.bronze", tabla=table, filas=n)
                _emit({"tipo": "hoja", "hoja": sheet, "estado": "ok",
                       "tabla": f"bronze.{table}", "filas": n})

        # ---- BRONZE landing (demás hojas) ----
        hojas_landing = 0
        for sheet in wb.sheetnames:
            if sheet in RAW_SHEETS: continue
            _emit({"tipo": "hoja", "hoja": sheet, "estado": "procesando"})
            n = land_landing(conn, wb[sheet], sheet, reporte_id)
            _log_ingesta(conn, reporte_id, sheet, "bronze.hoja_landing", n, n)
            hojas_landing += 1
            _emit({"tipo": "hoja", "hoja": sheet, "estado": "ok",
                   "tabla": "bronze.hoja_landing", "filas": n})
        filas["bronze.hoja_landing"] = hojas_landing
        log.info("ingesta.landing", hojas=hojas_landing)

        # ---- FACTS ECP (solo si raw) ----  [v4: emite el destino Core por hoja]
        if raw:
            n, sk = load_fact_dia(conn, wb["BDP_datos_dia"], reporte_id, (vice, socio, conc, tipo))
            _log_ingesta(conn, reporte_id, "BDP_datos_dia", "core.fact_produccion_dia_ecp", n+sk, n)
            filas["fact_produccion_dia_ecp"] = n
            log.info("ingesta.dia", filas=n, descartadas=sk)
            _emit({"tipo": "hoja", "hoja": "BDP_datos_dia", "estado": "ok",
                   "tabla": "fact_produccion_dia_ecp", "filas": n})

            n, sk = load_fact_mes(conn, wb["BDP_datos_mes"], reporte_id,
                                  (vice, socio, conc, tipo, esc, proc), emit=_emit)
            _log_ingesta(conn, reporte_id, "BDP_datos_mes", "core.fact_produccion_mes_ecp", n+sk, n)
            filas["fact_produccion_mes_ecp"] = n
            log.info("ingesta.mes", filas=n, descartadas=sk)
            _emit({"tipo": "hoja", "hoja": "BDP_datos_mes", "estado": "ok",
                   "tabla": "fact_produccion_mes_ecp", "filas": n})

            # pre-siembra de fuentes de programa que no estén aún (IDBDP nuevos) para no violar FK
            prg = wb["BDP_Programa"]
            it = prg.iter_rows(values_only=True); next(it, None)
            extra = {}
            for r in it:
                idb = num(r[10]) if r and len(r) > 10 else None
                if idb is not None: extra[int(idb)] = {"campo": s(r[7]), "grupo1": s(r[9]),
                                                        "gerencia": s(r[2]), "contrato": s(r[11])}
            upsert_fuentes(conn, extra, reporte_id)

            n, sk = load_fact_programa(conn, wb["BDP_Programa"], reporte_id, (vice, tipo))
            _log_ingesta(conn, reporte_id, "BDP_Programa", "core.fact_programa_ecp", n+sk, n)
            filas["fact_programa_ecp"] = n
            log.info("ingesta.programa", filas=n, descartadas=sk)
            _emit({"tipo": "hoja", "hoja": "BDP_Programa", "estado": "ok",
                   "tabla": "fact_programa_ecp", "filas": n})

        # ---- COMENTARIOS ----
        if "COMENTARIOS" in wb.sheetnames:
            n = load_comentarios(conn, wb["COMENTARIOS"], reporte_id, tipo)
            _log_ingesta(conn, reporte_id, "COMENTARIOS", "core.fact_comentarios_produccion", n, n)
            filas["fact_comentarios_produccion"] = n
            log.info("ingesta.comentarios", filas=n)
            _emit({"tipo": "hoja", "hoja": "COMENTARIOS", "estado": "ok",
                   "tabla": "fact_comentarios_produccion", "filas": n,
                   "reporte_id": reporte_id,
                   "tablas": [{"tabla_idx": 1, "tabla_label": "COMENTARIOS", "filas": n}]})

        # ---- FILIALES / POP / PROMEDIOS / CONFIG ----
        if "Producción filiales" in wb.sheetnames:
            n = load_filiales(conn, wb["Producción filiales"], reporte_id, emp, tipo, treg)
            _log_ingesta(conn, reporte_id, "Producción filiales", "core.fact_produccion_diaria", n, n)
            filas["fact_produccion_diaria"] = n
            log.info("ingesta.filiales", filas=n)
            _emit({"tipo": "hoja", "hoja": "Producción filiales", "estado": "ok",
                   "tabla": "fact_produccion_diaria", "filas": n})

        if "POP Filiales y Exploración" in wb.sheetnames:
            n = load_pop(conn, wb["POP Filiales y Exploración"], reporte_id, emp)
            _log_ingesta(conn, reporte_id, "POP Filiales y Exploración", "core.fact_plan_mensual", n, n)
            filas["fact_plan_mensual"] = n
            log.info("ingesta.pop", filas=n)
            _emit({"tipo": "hoja", "hoja": "POP Filiales y Exploración", "estado": "ok",
                   "tabla": "fact_plan_mensual", "filas": n})

        if "INICIO" in wb.sheetnames:
            n = load_promedios(conn, wb["INICIO"], reporte_id, emp, tipo)
            _log_ingesta(conn, reporte_id, "INICIO", "core.fact_promedio_validado", n, n)
            filas["fact_promedio_validado"] = n
            log.info("ingesta.promedios", filas=n)
            update_config_inicio(conn, wb["INICIO"], reporte_id)
            _emit({"tipo": "hoja", "hoja": "INICIO", "estado": "ok",
                   "tabla": "fact_promedio_validado", "filas": n})

        # ---- Hojas modeladas -> core.fact_tabla_hoja (registro escalable) ----
        for _res in load_tablas_hoja(conn, wb, reporte_id):
            filas[f"tabla_hoja::{_res['hoja']}"] = _res["total"]
            _log_ingesta(conn, reporte_id, _res["hoja"], "core.fact_tabla_hoja",
                         _res["total"], _res["total"])
            log.info("ingesta.tablahoja", hoja=_res["hoja"], total=_res["total"])
            _emit({"tipo": "hoja", "hoja": _res["hoja"], "estado": "ok",
                   "tabla": "fact_tabla_hoja", "filas": _res["total"],
                   "reporte_id": reporte_id, "tablas": _res["tablas"]})

    wb.close()
    tipo_archivo = "NEW" if raw else "STD"
    log.info("ingesta.ok", archivo=path.name, reporte_id=reporte_id, tipo=tipo_archivo, filas=filas)
    return ResultadoIngesta(
        archivo=path.name,
        reporte_id=reporte_id,
        tipo_archivo=tipo_archivo,
        tiene_raw=raw,
        filas_por_tabla=filas,
    )

# ================================================================ ORQUESTACIÓN PARA UI (jobs)
from app.features.ingesta.detector import nombres_de_hojas, tiene_raw

def _nombres_en_data() -> set[str]:
    """Set barato de nombres de .xlsm en data/ (NO abre los libros). Para validar solicitudes."""
    return {p.name for p in Path(get_settings().data_dir).rglob("*.xlsm")}

def listar_disponibles() -> list[dict]:
    """Lista los .xlsm de data/ con tipo NEW/STD (por hoja) y si ya fueron ingeridos.
    C2: to_date() devuelve dt.date y config_reporte.fecha_reporte es DATE -> la comparacion
    `fecha in ya` es date==date (correcta). NO cambiar a datetime."""
    base = Path(get_settings().data_dir)
    with get_engine().connect() as c:
        ya = {r[0] for r in c.execute(sa.text("SELECT fecha_reporte FROM core.config_reporte"))}
    out = []
    for p in _archivos_ordenados(base):
        m = re.search(r"(\d{8})", p.name)
        fecha = to_date(m.group(1)) if m else None
        raw = tiene_raw(nombres_de_hojas(p))   # C4: si el zip falla, set() => se reporta STD
        out.append({"nombre": p.name, "tipo": "NEW" if raw else "STD",
                    "fecha": fecha.isoformat() if fecha else None,
                    "ya_ingerido": fecha in ya})
    return out

def crear_job(nombres: list[str]) -> int:
    with get_engine().begin() as conn:
        return conn.execute(sa.text(
            "INSERT INTO core.ingesta_job (estado, total, archivos) "
            "VALUES ('PENDIENTE', :t, CAST(:a AS jsonb)) RETURNING job_id"),
            {"t": len(nombres), "a": json.dumps(nombres, ensure_ascii=False)}).scalar()

def _job_estado(job_id: int, estado: str, mensaje: str | None = None):
    with get_engine().begin() as c:
        c.execute(sa.text("UPDATE core.ingesta_job SET estado=:e, mensaje=:m, actualizado_at=now() "
                          "WHERE job_id=:j"), {"e": estado, "m": mensaje, "j": job_id})

def _job_progreso(job_id: int, procesados: int, errores: int, resultado: list[dict]):
    with get_engine().begin() as c:
        c.execute(sa.text("UPDATE core.ingesta_job SET procesados=:p, errores=:e, "
                          "resultado=CAST(:r AS jsonb), actualizado_at=now() WHERE job_id=:j"),
                  {"p": procesados, "e": errores,
                   "r": json.dumps(resultado, ensure_ascii=False), "j": job_id})

def procesar_job(job_id: int, paths: list[Path]):
    """Worker de BackgroundTasks: ingiere los archivos en orden, actualizando progreso por archivo."""
    _job_estado(job_id, "EN_PROCESO")
    resultado, procesados, errores = [], 0, 0
    try:
        for p in paths:
            try:
                r = ingerir_archivo(p)
                resultado.append({"archivo": p.name, "reporte_id": r.reporte_id, "tipo": r.tipo_archivo})
            except Exception as e:   # un archivo no debe tumbar el lote
                errores += 1
                resultado.append({"archivo": p.name, "error": str(e)})
                log.error("job.archivo.error", job=job_id, archivo=p.name, error=str(e))
            procesados += 1
            _job_progreso(job_id, procesados, errores, resultado)
        _job_estado(job_id, "COMPLETADO")
        log.info("job.completado", job=job_id, total=len(paths), errores=errores)
    except Exception as e:           # fallo inesperado a nivel job
        _job_estado(job_id, "ERROR", str(e))
        log.error("job.error", job=job_id, error=str(e))

def obtener_job(job_id: int) -> dict | None:
    with get_engine().connect() as c:
        r = c.execute(sa.text("SELECT * FROM core.ingesta_job WHERE job_id=:j"),
                      {"j": job_id}).mappings().first()
    return dict(r) if r else None

def listar_jobs(limite: int = 20) -> list[dict]:
    with get_engine().connect() as c:
        rows = c.execute(sa.text(
            "SELECT job_id, estado, total, procesados, errores, mensaje, creado_at, actualizado_at "
            "FROM core.ingesta_job ORDER BY job_id DESC LIMIT :l"), {"l": limite}).mappings().all()
    return [dict(r) for r in rows]
