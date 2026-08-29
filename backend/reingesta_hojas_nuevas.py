"""
Re-ingesta TARGETADA de las hojas nuevas + limpieza — para el servidor 139.
=============================================================================

QUÉ HACE
  Re-extrae SOLO 2 hojas (rápido, sin reprocesar los facts pesados) y las reemplaza
  en core.fact_tabla_hoja con la lógica DELETE+INSERT por hoja (idéntica al motor):
    · P50 Acumulado    → arregla la Tabla 2 contaminada + agrega RETO CORPORATIVO (Tablas 3/4).
    · REPORTE_PRESIDENT → tabla comparativa por producto (Real/Proy/P50/Compromiso, con BLANCOS).
  El DELETE+INSERT por hoja ES la limpieza: borra las filas viejas (incluida la basura
  'Días del año' / ',' / valores conflados de la Tabla 2 rota) y deja solo las correctas.

  NO reprocesa BDP/facts/comentarios → es mucho más liviano que `app.cli archivo`.
  NO toca ninguna otra hoja ni el resto de los datos de 139.

REQUISITOS EN 139
  1. Código ya desplegado (git pull) con los 2 extractores nuevos.
  2. Los .xlsm de los reportes a corregir, en una carpeta accesible (basta el MES en curso;
     no hace falta el histórico completo). Se emparejan con el reporte existente por la FECHA
     del nombre (YYYYMMDD...) contra core.config_reporte.fecha_reporte.
  3. El .env de 139 activo (get_engine() lee la conexión de INGESTA/Rep_Prod/.env).

USO  (desde INGESTA/Rep_Prod/backend, con el intérprete del proyecto)
  uv run python reingesta_hojas_nuevas.py <CARPETA_XLSM> [--dry-run]
    <CARPETA_XLSM>  carpeta con los .xlsm a corregir (ej. D:/reportes/Mayo_2026).
                    Si se omite, usa el data_dir configurado.
    --dry-run       muestra qué haría (filas por hoja) SIN escribir en la BD.

SEGURIDAD
  Idempotente (correrlo 2 veces deja el mismo resultado). Resiliente (un archivo con error
  no detiene el lote). Solo UPDATE de reportes YA existentes (si un .xlsm no tiene reporte en
  config_reporte, se omite con aviso — no crea reportes nuevos).
"""
import sys
import re
import json
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import sqlalchemy as sa
from openpyxl import load_workbook

from app.core.db import get_engine
from app.core.config import get_settings
from app.features.ingesta.services import _p50_acum_extract, _reporte_president_extract

# (regex de nombre de hoja, extractor, etiqueta legible)
TARGETS = [
    (re.compile(r"(?i)^P50 Acumulado"), _p50_acum_extract, "P50 Acumulado"),
    (re.compile(r"(?i)^REPORTE_PRESIDENT$"), _reporte_president_extract, "REPORTE_PRESIDENT"),
]
HOJAS = [t[2] for t in TARGETS]
CHUNK = 10_000

_INS = sa.text("""
    INSERT INTO core.fact_tabla_hoja (reporte_id, hoja, tabla_idx, tabla_label, dims, fecha, valor)
    VALUES (:r, :h, :idx, :label, CAST(:dims AS jsonb), :fecha, :valor)
""")


def _fecha_de_nombre(nombre):
    """YYYYMMDD del inicio del nombre -> 'YYYY-MM-DD' (o None)."""
    m = re.match(r"(\d{4})(\d{2})(\d{2})", nombre)
    return f"{m.group(1)}-{m.group(2)}-{m.group(3)}" if m else None


def _dedup(filas):
    """Última-gana por (tabla_idx, dims, fecha) — idéntico a load_tablas_hoja."""
    by = {}
    for f in filas:
        k = (f["tabla_idx"], json.dumps(f["dims"], sort_keys=True, ensure_ascii=False), f["fecha"])
        by[k] = f
    return list(by.values())


def _procesar_archivo(engine, path, dry):
    fecha = _fecha_de_nombre(path.name)
    if not fecha:
        return ("skip", f"nombre sin fecha YYYYMMDD: {path.name}")
    with engine.connect() as c:
        rid = c.execute(
            sa.text("SELECT reporte_id FROM core.config_reporte WHERE fecha_reporte = CAST(:f AS date)"),
            {"f": fecha}).scalar()
    if not rid:
        return ("skip", f"{fecha}: sin reporte en config_reporte (no ingerido) — se omite")

    wb = load_workbook(path, read_only=True, data_only=True)
    detalle = []
    try:
        for pat, extractor, etiqueta in TARGETS:
            hoja = next((sh for sh in wb.sheetnames if pat.match(sh)), None)
            if not hoja:
                detalle.append(f"{etiqueta}: hoja ausente")
                continue
            res = extractor(wb[hoja])
            filas = res.get("rows", []) if isinstance(res, dict) else res
            out = _dedup(filas)
            if dry:
                detalle.append(f"{hoja}: {len(out)} filas (dry-run, no escrito)")
                continue
            with engine.begin() as c:
                c.execute(sa.text("DELETE FROM core.fact_tabla_hoja WHERE reporte_id=:r AND hoja=:h"),
                          {"r": rid, "h": hoja})
                for i in range(0, len(out), CHUNK):
                    c.execute(_INS, [
                        {"r": rid, "h": hoja, "idx": f["tabla_idx"], "label": f["tabla_label"],
                         "dims": json.dumps(f["dims"], ensure_ascii=False),
                         "fecha": f["fecha"], "valor": f["valor"]}
                        for f in out[i:i + CHUNK]])
            detalle.append(f"{hoja}: {len(out)} filas reemplazadas")
    finally:
        wb.close()
    return ("ok", f"reporte {rid} ({fecha}) → " + " · ".join(detalle))


def _diagnostico_limpieza(engine):
    """Estado de limpieza global: reportes cuya P50 Acumulado sigue con la Tabla 2 corrupta
    (>48 filas = arrastra el bloque RETO/basura) y cuántos ya tienen las 4 tablas nuevas."""
    with engine.connect() as c:
        corruptos = c.execute(sa.text("""
            SELECT COUNT(*) FROM (
                SELECT reporte_id FROM core.fact_tabla_hoja
                WHERE hoja='P50 Acumulado' AND tabla_idx=2
                GROUP BY reporte_id HAVING COUNT(*) > 48
            ) q""")).scalar() or 0
        con_reto = c.execute(sa.text("""
            SELECT COUNT(DISTINCT reporte_id) FROM core.fact_tabla_hoja
            WHERE hoja='P50 Acumulado' AND tabla_idx=3""")).scalar() or 0
        con_pres = c.execute(sa.text("""
            SELECT COUNT(DISTINCT reporte_id) FROM core.fact_tabla_hoja
            WHERE hoja='REPORTE_PRESIDENT'""")).scalar() or 0
    print("\n--- Diagnóstico de limpieza (global) ---")
    print(f"  Reportes con RETO CORPORATIVO ya ingerido (P50 Acumulado, Tabla 3): {con_reto}")
    print(f"  Reportes con REPORTE_PRESIDENT ya ingerido: {con_pres}")
    print(f"  Reportes con P50 Acumulado AÚN corrupta (Tabla 2 > 48 filas): {corruptos}"
          + ("  ✅ limpio" if corruptos == 0 else "  ← re-procesar sus .xlsm para limpiarlos"))


def main():
    argv = sys.argv[1:]
    dry = "--dry-run" in argv
    pos = [a for a in argv if not a.startswith("--")]
    carpeta = Path(pos[0]) if pos else Path(get_settings().data_dir)

    if not carpeta.exists():
        print(f"ERROR: la carpeta no existe: {carpeta}")
        sys.exit(1)

    archivos = sorted(carpeta.rglob("*.xlsm"))
    if not archivos:
        print(f"No se encontraron .xlsm en: {carpeta}")
        sys.exit(1)

    print(f"Carpeta: {carpeta}")
    print(f"Archivos .xlsm: {len(archivos)}")
    print(f"Hojas objetivo: {', '.join(HOJAS)}")
    print(f"Modo: {'DRY-RUN (no escribe)' if dry else 'ESCRITURA'}\n")

    ok = skip = err = 0
    engine = get_engine()
    for i, path in enumerate(archivos, 1):
        try:
            estado, msg = _procesar_archivo(engine, path, dry)
        except Exception as e:  # un archivo no debe tumbar el lote
            estado, msg = "error", f"{path.name}: {type(e).__name__}: {e}"
        if estado == "ok":
            ok += 1
        elif estado == "skip":
            skip += 1
        else:
            err += 1
        print(f"[{i}/{len(archivos)}] {estado.upper():5} {msg}")

    print(f"\nRESUMEN: {ok} procesados · {skip} omitidos · {err} errores"
          + ("  (DRY-RUN)" if dry else ""))
    if not dry:
        _diagnostico_limpieza(engine)


if __name__ == "__main__":
    main()
